from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
import logging
from django.db.models import Q
from django.utils import timezone
from django.contrib import messages
from django.utils.translation import gettext as _
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required

from .models import Visit, Diagnosis, Room, EmergencyObservation, EmergencyMedication, VisitDiagnosis, VitalSigns
from .forms import (
    ExaminationForm, TriageForm, EmergencyExaminationForm,
    EmergencyAdmissionUpdateForm, EmergencyObservationForm,
    EmergencyMedicationForm, EmergencyDischargeForm
)
from .utils import log_visit_action

@login_required
@permission_required('medical_records.add_visit', raise_exception=True)
def emergency_direct_registration(request):
    """Fast-track registration specifically for Emergency Room direct arrivals."""
    from patients.models import Patient, DailyQueue
    from patients.forms import PatientRegistrationForm
    from django.db.models import Q
    
    query = request.GET.get('q', '').strip()
    patients = []
    if query:
        patients = Patient.objects.filter(
            Q(full_name__icontains=query) | 
            Q(patient_id__icontains=query) |
            Q(phone_number__icontains=query)
        )[:10]

    if request.method == 'POST' and 'register_new' in request.POST:
        form = PatientRegistrationForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            from django.db.models import Max
            year = timezone.localdate().year
            prefix = f"MD{year}"
            last_patient = Patient.objects.filter(patient_id__startswith=prefix).aggregate(
                max_id=Max('patient_id')
            )
            if last_patient['max_id']:
                last_num = int(last_patient['max_id'][-4:])
                new_num = last_num + 1
            else:
                new_num = 1
            patient.patient_id = f"{prefix}{new_num:04d}"
            patient.save()
            
            # Auto create visit to Emergency with arrival details
            arrival_mode = request.POST.get('arrival_mode', 'WALK_IN')
            brought_by = request.POST.get('brought_by_name', '')
            emergency_room = Room.objects.filter(code='IGD').first() or Room.objects.filter(code='EMERGENCY').first()
            queue, created = DailyQueue.objects.get_or_create(date=timezone.localdate(), department='General')
            
            visit = Visit.objects.create(
                patient=patient,
                current_room=emergency_room,
                status='IP',
                queue_number=queue.get_next_number(),
                checked_in_by=request.user,
                doctor=request.user,
                arrival_mode=arrival_mode,
                brought_by_name=brought_by,
                patient_type='FOUN'
            )
            log_visit_action(visit, 'CHECK_IN', request.user, room=emergency_room)
            messages.success(request, _("Patient %(name)s registered and admitted to Emergency.") % {'name': patient.full_name})
            return redirect('perform_examination', visit_uuid=visit.uuid)
    
    elif request.method == 'POST' and 'select_patient' in request.POST:
        patient_uuid = request.POST.get('patient_uuid')
        patient = get_object_or_404(Patient, uuid=patient_uuid)
        
        # Auto create visit to Emergency
        emergency_room = Room.objects.filter(code='IGD').first() or Room.objects.filter(code='EMERGENCY').first()
        queue, created = DailyQueue.objects.get_or_create(date=timezone.localdate(), department='General')
        
        visit = Visit.objects.create(
            patient=patient,
            current_room=emergency_room,
            status='IP',
            queue_number=queue.get_next_number(),
            checked_in_by=request.user,
            doctor=request.user,
            arrival_mode=request.POST.get('arrival_mode', 'WALK_IN'),
            brought_by_name=request.POST.get('brought_by_name', ''),
            companion_name=request.POST.get('companion_name', ''),
            patient_type='TUAN'
        )
        log_visit_action(visit, 'CHECK_IN', request.user, room=emergency_room)
        messages.success(request, _("Visit started for %(name)s in Emergency.") % {'name': patient.full_name})
        return redirect('perform_examination', visit_uuid=visit.uuid)

    next_id = Patient.generate_next_id()
    registration_form = PatientRegistrationForm(initial={'patient_id': next_id})
    return render(request, 'medical_records/emergency_direct_registration.html', {
        'patients': patients,
        'query': query,
        'registration_form': registration_form,
    })


@login_required
@permission_required('medical_records.view_menu_medical_records', raise_exception=True)
def triage_dashboard(request):
    """Room 2: Triage Dashboard for nurses."""
    query = request.GET.get('q', '')
    today = timezone.localdate()
    
    # Base filter for Room 2 waiting:
    # - Include 'In Progress' visits from any day (to allow follow-up)
    # - Include 'Scheduled' visits ONLY for today (to avoid no-show clutter)
    waiting_visits = Visit.objects.visible_to(request.user).filter(
        current_room__code__in=['TRIAGE', 'ROOM_2'],
    ).filter(
        Q(status='IP') | Q(status='SCH', visit_date__date=today)
    ).select_related('patient', 'current_room')
    
    # If search used
    search_results = None
    if query:
        search_results = Visit.objects.filter(
            Q(patient__full_name__icontains=query) | Q(patient__patient_id__icontains=query),
            visit_date__date=today
        ).select_related('patient', 'current_room').order_by('queue_number')[:50]
    
    waiting_visits = waiting_visits.order_by('queue_number')
    
    completed_today = Visit.objects.visible_to(request.user).filter(
        visit_date__date=today,
        current_room__order__gt=2
    ).exclude(current_room__code__in=['IGD', 'EMERGENCY']).select_related('patient', 'current_room').order_by('-visit_date')[:100]
    
    return render(request, 'medical_records/triage_dashboard.html', {
        'waiting_visits': waiting_visits,
        'completed_today': completed_today,
        'search_results': search_results,
        'query': query,
        'stats': {
            'waiting': waiting_visits.count(),
            'completed': completed_today.count(),
        }
    })

@login_required
@permission_required('medical_records.change_visit', raise_exception=True)
def triage_input(request, visit_uuid):
    """Room 2: Triage Input Form."""
    from .forms import VitalSignsForm
    from django.shortcuts import get_object_or_404
    from .utils import log_visit_action
    visit = get_object_or_404(Visit, uuid=visit_uuid)
    vitals, _c = VitalSigns.objects.get_or_create(visit=visit)
    
    if request.method == 'POST':
        form = TriageForm(request.POST, instance=visit)
        vitals_form = VitalSignsForm(request.POST, instance=vitals, triage_level=request.POST.get('triage_level'))
        
        if form.is_valid() and vitals_form.is_valid():
            try:
                v = form.save(commit=False)
                v.triage_nurse = request.user
                v.save()
                
                v_signs = vitals_form.save(commit=False)
                v_signs.visit = v
                v_signs.save()
                
                log_visit_action(v, 'TRIAGE', request.user, room=v.current_room)
                
                room_name = v.current_room.name if v.current_room else _("Doctor")
                messages.success(request, _("Triage for %(name)s completed. Patient directed to %(room)s.") % {
                    'name': v.patient.full_name,
                    'room': room_name
                })
                return redirect('triage_dashboard')
            except Exception as e:
                import logging
                logging.getLogger(__name__).error(f"Error saving triage: {str(e)}", exc_info=True)
                messages.error(request, _("A system error occurred while saving triage: %s") % str(e))
                return redirect('triage_input', visit_uuid=v.uuid)
    else:
        form = TriageForm(instance=visit)
        vitals_form = VitalSignsForm(instance=vitals)
        if visit.current_room and visit.current_room.code not in ['TRIAGE', 'ROOM_2']:
             messages.warning(request, _("This patient has already been triaged."))
             
    return render(request, 'medical_records/triage_form.html', {
        'visit': visit,
        'form': form,
        'vitals_form': vitals_form,
    })

@login_required
@permission_required('medical_records.view_menu_medical_records', raise_exception=True)
def doctor_dashboard(request):
    today = timezone.localdate()
    room_filter = request.GET.get('room', '')  # e.g. ?room=KIA
    search_query = request.GET.get('q', '').strip()

    # Rooms that should appear in the general OPD dashboard
    general_rooms = ['DOKTER', 'ROOM_3', 'ROOM_4', 'ROOM_5', 'ROOM_6']
    
    # Specialized rooms that require explicit selection
    specialist_rooms = ['KIA', 'HIV', 'TB', 'DENTAL', 'NUTRISI', 'USG']
    
    all_clinical_rooms = general_rooms + specialist_rooms

    # Role-based room default for non-admins
    if not room_filter and not request.user.is_superuser:
        if hasattr(request.user, 'staff_profile'):
            dept_code = request.user.staff_profile.department.code.upper()
            if dept_code in specialist_rooms:
                room_filter = dept_code
            elif dept_code in general_rooms:
                room_filter = dept_code

    if room_filter and room_filter in all_clinical_rooms:
        # Show only patients in that specific specialist room
        room_codes = [room_filter]
        page_title = Room.objects.filter(code=room_filter).values_list('name', flat=True).first() or room_filter
    else:
        # Default: show ONLY general clinical rooms
        room_codes = general_rooms
        page_title = _("General Consultation")

    # Show pending/in-progress patients:
    # - Include 'In Progress' (IP) from any date to ensure continuity.
    # - Include 'Scheduled' (SCH) only for today.
    waiting_visits_qs = Visit.objects.visible_to(request.user).filter(
        Q(current_room__code__in=room_codes) |
        Q(doctor=request.user, status='IP')
    ).filter(
        Q(status='IP') | Q(status='SCH', visit_date__date=today)
    ).distinct()

    if search_query:
        waiting_visits_qs = waiting_visits_qs.filter(
            Q(patient__first_name__icontains=search_query) |
            Q(patient__last_name__icontains=search_query) |
            Q(patient__patient_id__icontains=search_query)
        )

    # Split visits into New (Today) and Continued (Leftover from previous days)
    new_visits = waiting_visits_qs.filter(visit_date__date=today).select_related('patient', 'current_room', 'doctor').order_by('queue_number')
    continued_visits_list = waiting_visits_qs.filter(visit_date__date__lt=today).select_related('patient', 'current_room', 'doctor').order_by('visit_date', 'queue_number')

    completed_visits = Visit.objects.visible_to(request.user).filter(
        current_room__code__in=room_codes,
        visit_date__date=today,
        status='COM'
    ).select_related('patient', 'current_room', 'doctor').order_by('-visit_date')[:100]

    room_counts = Visit.objects.visible_to(request.user).filter(
        current_room__code__in=room_codes,
        visit_date__date=today
    )

    # Cumulative totals for specific departments (e.g. HIV)
    total_patients_overall = 0
    if room_filter == 'HIV':
        from patients.models import Patient
        total_patients_overall = Patient.objects.filter(is_hiv_patient=True).count()

    return render(request, 'medical_records/doctor_dashboard.html', {
        'waiting_visits': new_visits,
        'continued_visits': continued_visits_list,
        'completed_visits': completed_visits,
        'room_filter': room_filter,
        'page_title': page_title,
        'stats': {
            'total_today': room_counts.count(),
            'new_pending': new_visits.count(),
            'continued_pending': continued_visits_list.count(),
            'completed_today': room_counts.filter(status='COM').count(),
            'total_overall': total_patients_overall,
        }
    })

@login_required
@permission_required('medical_records.view_menu_medical_records', raise_exception=True)
def emergency_dashboard(request):
    """Room: EMERGENCY Dashboard for critical care."""
    today = timezone.localdate()
    
    # Filter for active Emergency cases:
    # 1. Patients physically in IGD/EMERGENCY rooms
    # 2. Patients in auxiliary rooms (Lab, Rad, etc.) who have ER-specific vitals recorded (indicating they are from ER)
    waiting_visits_qs = Visit.objects.visible_to(request.user).filter(
        status__in=['SCH', 'IP']
    ).filter(
        Q(current_room__code__in=['IGD', 'EMERGENCY']) | 
        Q(source='IGD')
    ).select_related('patient', 'current_room').distinct()
    
    waiting_visits = waiting_visits_qs.order_by('queue_number')
    
    # Calculate observation duration for each visit
    for v in waiting_visits:
        v.observation_hours = (timezone.now() - v.visit_date).total_seconds() / 3600
        v.needs_referral = v.observation_hours > 24
    
    completed_today = Visit.objects.visible_to(request.user).filter(
        visit_date__date=today,
        current_room__code__in=['IGD', 'EMERGENCY'],
        status='COM'
    ).select_related('patient', 'current_room').order_by('-visit_date')[:100]
    
    # Emergency Stats
    room_counts = Visit.objects.visible_to(request.user).filter(
        current_room__code__in=['IGD', 'EMERGENCY'],
        visit_date__date=today
    )

    # Role detection
    is_nurse = False
    if hasattr(request.user, 'staff_profile'):
        is_nurse = request.user.staff_profile.category.name.upper() == 'PARAMEDIS'

    return render(request, 'medical_records/emergency_dashboard.html', {
        'waiting_visits': waiting_visits,
        'completed_today': completed_today,
        'stats': {
            'active_cases': waiting_visits.count(),
            'total_today': room_counts.count(),
            'discharged_today': room_counts.filter(status='COM').count(),
        },
        'is_nurse': is_nurse,
    })


@login_required
@permission_required('medical_records.view_menu_medical_records', raise_exception=True)
def emergency_triage_dashboard(request):
    """Dashboard specifically for ER Vitals/Triage check."""
    today = timezone.localdate()
    # Show patients in ER queue who haven't had their 'Emergency Re-check' vitals yet
    # Or just all active ER cases for triage
    waiting_visits = Visit.objects.visible_to(request.user).filter(
        current_room__code__in=['IGD', 'EMERGENCY'],
        status__in=['SCH', 'IP']
    ).select_related('patient', 'current_room').order_by('queue_number')
    
    # Simple check: if vitals record is missing or bp_sys is null, they definitely need triage
    for v in waiting_visits:
        v.needs_vitals = v.vitals is None or v.vitals.bp_sys is None
        
    return render(request, 'medical_records/emergency_triage_dashboard.html', {
        'waiting_visits': waiting_visits,
        'stats': {
            'waiting': waiting_visits.count(),
        }
    })


@login_required
@permission_required('medical_records.change_visit', raise_exception=True)
def emergency_triage_input(request, visit_uuid):
    """Focused triage/vitals input for Emergency Room."""
    visit = get_object_or_404(Visit, uuid=visit_uuid)
    from .forms import EmergencyExaminationForm, VitalSignsForm
    
    # Get or create vitals
    vitals, _c = VitalSigns.objects.get_or_create(visit=visit)
    
    if request.method == 'POST':
        form = EmergencyExaminationForm(request.POST, instance=visit)
        vitals_form = VitalSignsForm(request.POST, instance=vitals, triage_level=request.POST.get('triage_level'))
        
        if form.is_valid() and vitals_form.is_valid():
            v = form.save(commit=False)
            v.triage_nurse = request.user
            v.status = 'IP' 
            v.save()
            form.save_m2m()
            
            v_signs = vitals_form.save(commit=False)
            v_signs.visit = v
            v_signs.save()
            
            from .utils import log_visit_action
            log_visit_action(v, 'TRIAGE', request.user, room=v.current_room)
            messages.success(request, _("ER Vitals for %(name)s recorded.") % {'name': v.patient.full_name})
            return redirect('emergency_triage_dashboard')
    else:
        form = EmergencyExaminationForm(instance=visit)
        vitals_form = VitalSignsForm(instance=vitals)
        
    return render(request, 'medical_records/emergency_triage_form.html', {
        'visit': visit,
        'form': form,
        'vitals_form': vitals_form,
    })


@login_required
@permission_required('medical_records.change_visit', raise_exception=True)
def record_emergency_observation(request, visit_uuid):
    """Record hourly/periodic vitals in ER."""
    visit = get_object_or_404(Visit, uuid=visit_uuid)
    if request.method == 'POST':
        form = EmergencyObservationForm(request.POST)
        if form.is_valid():
            obs = form.save(commit=False)
            obs.visit = visit
            if not obs.checked_by:
                obs.checked_by = request.user
            obs.save()
            messages.success(request, _("Observation recorded for %(name)s.") % {'name': visit.patient.full_name})
            return redirect('perform_examination', visit_uuid=visit.uuid)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Observation Error - {field}: {error}")
    return redirect('perform_examination', visit_uuid=visit.uuid)


@login_required
@permission_required('medical_records.change_visit', raise_exception=True)
def administer_emergency_medication(request, visit_uuid):
    """Administer medication and deduct from Pharmacy stock immediately."""
    from django.db import transaction
    visit = get_object_or_404(Visit, uuid=visit_uuid)
    
    if request.method == 'POST':
        form = EmergencyMedicationForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                med_admin = form.save(commit=False)
                med_admin.visit = visit
                
                if not med_admin.given_by:
                    med_admin.given_by = request.user
                
                if not med_admin.ordered_by:
                    med_admin.ordered_by = visit.doctor
                
                # STOCK DEDUCTION LOGIC (FEFO)
                medicine = med_admin.medicine
                if medicine.total_stock >= med_admin.quantity:
                    remaining_to_deduct = med_admin.quantity
                    batches = medicine.stock_batches.filter(quantity_remaining__gt=0).order_by('expiry_date', 'purchase_date')
                    
                    for batch in batches:
                        if remaining_to_deduct <= 0:
                            break
                        deduct_qty = min(batch.quantity_remaining, remaining_to_deduct)
                        batch.quantity_remaining -= deduct_qty
                        batch.save()
                        remaining_to_deduct -= deduct_qty
                    
                    med_admin.save()
                    
                    from .utils import log_visit_action
                    log_visit_action(visit, 'PRESCRIPTION', request.user, 
                                     notes=f"Administered {med_admin.quantity} {medicine.name} ({med_admin.get_admin_type_display()})")
                    
                    messages.success(request, _("%(qty)d x %(med)s administered and stock updated.") % {
                        'qty': med_admin.quantity, 'med': medicine.name
                    })
                else:
                    messages.error(request, _("Insufficient stock for %(med)s. Current: %(stock)d") % {
                        'med': medicine.name, 'stock': medicine.total_stock
                    })
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Medication Error - {field}: {error}")
                    
    return redirect('perform_examination', visit_uuid=visit.uuid)

@login_required
@permission_required('medical_records.view_menu_medical_records', raise_exception=True)
def staff_performance_report(request):
    """Report on staff activity and patient outcomes (Success vs Referral)."""
    from django.db.models import Count, Q
    
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    today = timezone.localdate()
    if not date_from:
        date_from = today.replace(day=1).isoformat() # Beginning of month
    if not date_to:
        date_to = today.isoformat()
        
    # Get all visits in range
    visits = Visit.objects.visible_to(request.user).filter(
        visit_date__date__gte=date_from,
        visit_date__date__lte=date_to
    ).select_related('patient', 'doctor', 'triage_nurse', 'checked_in_by', 'current_room')
    
    # Global Aggregates
    total_count = visits.count()
    completed_clinic = visits.filter(status='COM', refer_to_central=False).count()
    referred_count = visits.filter(refer_to_central=True).count()
    
    # Staff Metrics - Doctors
    doctor_stats = visits.values(
        'doctor__username', 'doctor__first_name', 'doctor__last_name'
    ).annotate(
        total=Count('uuid'),
        completed=Count('uuid', filter=Q(status='COM', refer_to_central=False)),
        referred=Count('uuid', filter=Q(refer_to_central=True)),
        in_progress=Count('uuid', filter=Q(status='IP')),
    ).exclude(doctor__isnull=True).order_by('-total')
    
    # Staff Metrics - Nurses (Triage)
    nurse_stats = visits.values(
        'triage_nurse__username', 'triage_nurse__first_name', 'triage_nurse__last_name'
    ).annotate(
        total=Count('uuid')
    ).exclude(triage_nurse__isnull=True).order_by('-total')
    
    # Staff Metrics - Reception (Check-in)
    reception_stats = visits.values(
        'checked_in_by__username', 'checked_in_by__first_name', 'checked_in_by__last_name'
    ).annotate(
        total=Count('uuid')
    ).exclude(checked_in_by__isnull=True).order_by('-total')

    return render(request, 'medical_records/staff_performance_report.html', {
        'date_from': date_from,
        'date_to': date_to,
        'total_count': total_count,
        'completed_clinic': completed_clinic,
        'referred_count': referred_count,
        'doctor_stats': doctor_stats,
        'nurse_stats': nurse_stats,
        'reception_stats': reception_stats,
        'recent_visits': visits.order_by('-visit_date')[:100],
    })

@login_required
@permission_required('medical_records.view_menu_medical_records', raise_exception=True)
def disease_statistics_report(request):
    """
    Statistics report grouped by hierarchical diagnoses.
    Shows totals for parent diagnoses and breakdowns for sub-types.
    """
    from django.db.models import Count, Q
    from django.utils import timezone
    
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    today = timezone.localdate()
    if not date_from:
        date_from = today.replace(day=1).isoformat()
    if not date_to:
        date_to = today.isoformat()
        
    visit_diagnoses = VisitDiagnosis.objects.filter(
        visit__visit_date__date__gte=date_from,
        visit__visit_date__date__lte=date_to
    ).select_related('diagnosis', 'visit')
    
    # Get all root diagnoses (those without a parent)
    roots = Diagnosis.objects.filter(parent__isnull=True).order_by('code')
    
    report_data = []
    
    for root in roots:
        # Get all related IDs for this branch (root + children recursive)
        related_ids = root.get_related_ids()
        
        root_count = visit_diagnoses.filter(diagnosis_id=root.id).count()
        child_total = visit_diagnoses.filter(diagnosis_id__in=related_ids).exclude(diagnosis_id=root.id).count()
        total_count = root_count + child_total
        
        if total_count > 0:
            # Breakdown of sub-types
            breakdown = visit_diagnoses.filter(diagnosis_id__in=related_ids).exclude(diagnosis_id=root.id).values(
                'diagnosis__code', 'diagnosis__name'
            ).annotate(count=Count('uuid')).order_by('-count')
            
            report_data.append({
                'root': root,
                'total': total_count,
                'root_only': root_count,
                'breakdown': breakdown
            })
            
    return render(request, 'medical_records/disease_report.html', {
        'date_from': date_from,
        'date_to': date_to,
        'report_data': report_data,
    })

@login_required
@permission_required('medical_records.change_visit', raise_exception=True)
def perform_examination(request, visit_uuid):
    from .forms import ExaminationForm, EmergencyExaminationForm, VitalSignsForm
    visit = get_object_or_404(Visit, uuid=visit_uuid)
    
    # Get or create vitals
    vitals, _c = VitalSigns.objects.get_or_create(visit=visit)
    
    # Only show emergency forms/features if the patient is actually in the Emergency room (IGD/EMERGENCY).
    is_emergency = visit.current_room and visit.current_room.code in ['IGD', 'EMERGENCY']
    
    # Role-based detection
    is_nurse = False
    staff_profile = getattr(request.user, 'staff_profile', None)
    if staff_profile and staff_profile.category:
        is_nurse = staff_profile.category.name.upper() == 'PARAMEDIS'
    
    FormClass = EmergencyExaminationForm if is_emergency else ExaminationForm


    
    if request.method == 'POST':
        # --- Handle Specialized ER Modals ---
        action = request.POST.get('action')
        
        if action == 'update_admission' and is_emergency:
            adm_form = EmergencyAdmissionUpdateForm(request.POST, instance=visit)
            if adm_form.is_valid():
                adm_form.save()
                messages.success(request, _("Admission details updated."))
                return redirect('perform_examination', visit_uuid=visit.uuid)

        if action == 'discharge_patient' and is_emergency:
            dis_form = EmergencyDischargeForm(request.POST, instance=visit)
            if dis_form.is_valid():
                v = dis_form.save(commit=False)
                v.status = 'COM'
                v.discharge_datetime = timezone.now()
                v.save()
                
                if v.follow_up_date:
                    from appointments.models import Appointment
                    import datetime

                    auto_notes = v.follow_up_notes if v.follow_up_notes else _("Automated follow-up from Emergency Room")
                    
                    # PRIVACY POLICY (Option B):
                    # If the patient is an HIV patient, route the follow-up appointment
                    # to the HIV department room — NOT the IGD room.
                    # This ensures follow-up appointments for HIV patients never appear
                    # in the general appointment calendar.
                    if v.patient.is_hiv_patient:
                        hiv_room = Room.objects.filter(code='HIV').first()
                        appt_room = hiv_room if hiv_room else v.current_room
                        appt_reason = f"[HIV Follow-Up] {auto_notes}"
                    else:
                        appt_room = v.current_room
                        appt_reason = f"Follow-Up: {auto_notes}"

                    Appointment.objects.get_or_create(
                        patient=v.patient,
                        appointment_date=v.follow_up_date,
                        doctor=v.doctor,
                        defaults={
                            'department': appt_room,
                            'appointment_time': datetime.time(9, 0),
                            'reason': appt_reason,
                            'created_by': request.user,
                        }
                    )

                log_visit_action(v, 'COMPLETED', request.user)
                
                if v.patient.is_hiv_patient and v.follow_up_date:
                    messages.success(request, _(
                        "Patient discharged from Emergency. Follow-up appointment has been registered in the HIV Department schedule (%(date)s)."
                    ) % {'date': v.follow_up_date.strftime('%d %b %Y')})
                else:
                    messages.success(request, _("Patient discharged from Emergency."))
                    
                return redirect('emergency_dashboard')

        # --- Standard Main Form Handling ---
        form = FormClass(request.POST, instance=visit)
        vitals_form = VitalSignsForm(request.POST, instance=vitals, triage_level=request.POST.get('triage_level'))

        if form.is_valid() and vitals_form.is_valid():
            try:
                # Process multiple referrals
                referral_rooms = list(form.cleaned_data.get('referral_rooms', []))
                original_room = visit.current_room

                # Shared Referral Logic for both OPD and IGD
                if referral_rooms:
                    new_room = referral_rooms[0]
                    if new_room != original_room:
                        from patients.models import DailyQueue
                        
                        # Update current room
                        visit.current_room = new_room
                        
                        # Assign a NEW queue number for the target department
                        today = timezone.localdate()
                        queue, created = DailyQueue.objects.get_or_create(
                            date=today, 
                            department=new_room.code
                        )
                        visit.queue_number = queue.get_next_number()
                        
                        # Set status to In Progress for the next room
                        visit.status = 'IP'

                if is_nurse:
                    # Nurses can only update specific fields (Vitals, Lab Request)
                    old_instance = Visit.objects.get(pk=visit.pk)
                    
                    visit = form.save(commit=False)
                    # Restore sensitive fields if necessary, but keep the new room/queue
                    visit.clinical_notes = old_instance.clinical_notes
                    visit.save()
                    form.save_m2m()
                else:
                    # Doctors have full control
                    visit = form.save(commit=False)
                    visit.doctor = request.user
                    visit.save()
                    form.save_m2m()

                    # --- Handle Diagnoses (VisitDiagnosis) ---
                    primary_diag = form.cleaned_data.get('diagnosis')
                    secondary_diags = form.cleaned_data.get('secondary_diagnoses', [])

                    # Clear existing and re-create (Sync approach)
                    VisitDiagnosis.objects.filter(visit=visit).delete()
                    
                    if primary_diag:
                        VisitDiagnosis.objects.create(visit=visit, diagnosis=primary_diag, is_primary=True)
                    
                    for sd in secondary_diags:
                        if sd != primary_diag:
                            VisitDiagnosis.objects.create(visit=visit, diagnosis=sd, is_primary=False)

                # Save vitals
                v_signs = vitals_form.save(commit=False)
                v_signs.visit = visit
                v_signs.save()

                # --- Handle Follow-up Appointment ---
                if visit.follow_up_date:
                    from appointments.models import Appointment
                    import datetime
                    auto_notes = visit.follow_up_notes if visit.follow_up_notes else "Automated follow-up"
                    Appointment.objects.get_or_create(
                        patient=visit.patient,
                        appointment_date=visit.follow_up_date,
                        doctor=visit.doctor,
                        defaults={
                            'department': visit.current_room,
                            'appointment_time': datetime.time(9, 0),
                            'reason': f"Follow-Up / Automated: {auto_notes}",
                            'created_by': request.user
                        }
                    )

                # --- Allergy Sync Logic ---
                allergy_str = form.cleaned_data.get('allergy_noted')
                if allergy_str:
                    from patients.models import PatientAllergy
                    PatientAllergy.objects.get_or_create(
                        patient=visit.patient,
                        allergen=allergy_str,
                        defaults={'reaction': _('Noted during ER visit')}
                    )
                    messages.info(request, _("Allergy '%s' added to patient record.") % allergy_str)

                # --- Unified Diagnostic Referral Logic ---
                source_tag = 'IGD' if is_emergency else 'OPD'
                
                # Universal source tracking on the Visit model
                if visit.source != source_tag:
                    visit.source = source_tag
                    visit.save()

                referred_to_lab = visit.lab_cbc
                referred_to_rad = False
                referred_to_pharm = getattr(visit, 'pharmacy_requested', False)
                referred_to_patho = False

                for room in referral_rooms:
                    room_code_upper = room.code.upper() if room.code else ''
                    room_name_upper = room.name.upper() if room.name else ''
                    
                    if 'LAB' in room_name_upper or 'LAB' in room_code_upper or room_code_upper == 'ROOM_7':
                        referred_to_lab = True
                    elif 'RADIOLOGY' in room_code_upper or 'RAD' in room_code_upper:
                        referred_to_rad = True
                    elif 'PHARM' in room_name_upper or 'PHARM' in room_code_upper or room_code_upper == 'ROOM_8':
                        referred_to_pharm = True
                    elif 'PATHOLOGY' in room_code_upper or 'PATHO' in room_code_upper:
                        referred_to_patho = True

                # 1. Laboratory Request
                if referred_to_lab:
                    from laboratory.models import LabRequest, LabTest
                    lab_req, created = LabRequest.objects.get_or_create(
                        visit=visit,
                        defaults={
                            'requesting_physician': request.user,
                            'source': source_tag
                        }
                    )
                    if not created:
                        lab_req.source = source_tag
                        lab_req.requesting_physician = request.user
                        lab_req.save()

                    if visit.lab_cbc:
                        # Use .filter().first() instead of .get() to avoid MultipleObjectsReturned/DoesNotExist crashes
                        cbc_test = LabTest.objects.filter(name__iexact='CBC').first()
                        if cbc_test:
                            lab_req.tests.add(cbc_test)
                
                # 2. Pharmacy / Prescription
                if referred_to_pharm:
                    from pharmacy.models import Prescription
                    presc, created = Prescription.objects.get_or_create(
                        visit=visit,
                        defaults={
                            'doctor': request.user,
                            'source': source_tag
                        }
                    )
                    if not created:
                        presc.source = source_tag
                        presc.doctor = request.user
                        presc.save()

                # 3. Radiology Request
                if referred_to_rad:
                    from radiology.models import RadiologyRequest
                    rad_req, created = RadiologyRequest.objects.get_or_create(
                        visit=visit,
                        defaults={
                            'requesting_physician': request.user,
                            'source': source_tag
                        }
                    )
                    if not created:
                        rad_req.requesting_physician = request.user
                        rad_req.save()

                # 4. Pathology Request
                if referred_to_patho:
                    from pathology.models import PathologyRequest
                    PathologyRequest.objects.get_or_create(
                        visit=visit,
                        defaults={
                            'requesting_physician': request.user,
                            'source': source_tag
                        }
                    )

                log_visit_action(visit, 'COMPLETED' if visit.status == 'COM' else 'EXAMINATION', request.user, room=visit.current_room)

            except Exception as e:
                import logging
                logging.getLogger(__name__).error(f"Error saving examination: {str(e)}", exc_info=True)
                messages.error(request, _("A system error occurred while saving: %s") % str(e))
                return redirect('perform_examination', visit_uuid=visit.uuid)

            if visit.refer_to_central:
                log_visit_action(visit, 'REFERRED', request.user)

            if is_emergency:
                dept_names = ", ".join([r.name for r in referral_rooms]) if referral_rooms else _("None")
                messages.success(request, _("Data examination for %(name)s saved. Referral sent to %(dept)s. Patient remains in Emergency queue.") % {
                    'name': visit.patient.full_name,
                    'dept': dept_names,
                })
            else:
                messages.success(request, _("Examination for %(name)s saved. Patient referred to %(room)s.") % {
                    'name': visit.patient.full_name,
                    'room': visit.current_room.name if visit.current_room else _("None")
                })

            if visit.status == 'COM' or (referral_rooms and visit.current_room != original_room) or is_emergency:
                if is_emergency:
                    return redirect('emergency_dashboard')
                return redirect('doctor_dashboard')
            
            return redirect('perform_examination', visit_uuid=visit.uuid)

    else:
        # Load existing diagnoses into the form
        primary_diag = VisitDiagnosis.objects.filter(visit=visit, is_primary=True).first()
        secondary_diags = Diagnosis.objects.filter(visit_records__visit=visit, visit_records__is_primary=False)
        
        form = FormClass(instance=visit, initial={
            'diagnosis': primary_diag.diagnosis if primary_diag else None,
            'secondary_diagnoses': secondary_diags
        })
        vitals_form = VitalSignsForm(instance=vitals)
    
    enrich_visit_lab_results(visit)

    # Prepare extra forms for Emergency cases
    context = {
        'visit': visit,
        'form': form,
        'vitals_form': vitals_form,
        'is_emergency': is_emergency,
        'is_nurse': is_nurse,
    }

    if is_emergency:
        context['obs_form'] = EmergencyObservationForm(initial={'check_time': timezone.now(), 'checked_by': request.user})
        context['med_form'] = EmergencyMedicationForm(initial={'ordered_by': visit.doctor, 'given_by': request.user})
        context['adm_form'] = EmergencyAdmissionUpdateForm(instance=visit)
        context['dis_form'] = EmergencyDischargeForm(instance=visit)
        context['emergency_observations'] = visit.observations.all().order_by('-check_time')
        context['administered_medications'] = visit.emergency_medications.all().order_by('-given_at')

    # --- Enrich Diagnostic Data (Lab, Radiology, Images) ---
    enrich_visit_diagnostic_data(visit, context)

    return render(request, 'medical_records/examination.html', context)


def enrich_visit_diagnostic_data(visit, context):
    """Aggregates all diagnostic results, reports, and images for display."""
    # 1. Lab Results Enrichment
    lab_results = []
    # Note: 'lab_results' relation is obsolete in new schema, we now use visit.lab_request.result.result_data
    # This block is kept for template compatibility but initialized safely as empty
    context['lab_results'] = lab_results

    
    # 2. Radiology Results & Images
    radiology_data = {
        'request': None,
        'result': None,
        'images': []
    }
    
    if hasattr(visit, 'radiology_request'):
        rad_req = visit.radiology_request
        radiology_data['request'] = rad_req
        if hasattr(rad_req, 'result'):
            rad_res = rad_req.result
            radiology_data['result'] = rad_res
            # Fetch all attachments (gallery)
            radiology_data['images'] = list(rad_res.attachments.all())
            
    context['radiology_data'] = radiology_data


def enrich_visit_lab_results(visit):
    from laboratory.views import CBC_RANGES, SEROLOGY_PARAMS, BIOCHEMISTRY_PARAMS, URINALYSIS_PARAMS, MICROSCOPY_PARAMS, MICROBIOLOGY_PARAMS
    try:
        lab_req = getattr(visit, 'lab_request', None)
        if lab_req and lab_req.status == 'COMPLETED':
            res = lab_req.result
            if not res or not res.result_data:
                return

            # 1. CBC Handling (Legacy or simple CBC only)
            cbc_data = res.result_data.get('cbc')
            if isinstance(cbc_data, dict):
                cat = res.result_data.get('category')
                ranges = CBC_RANGES.get(cat, {})
                cbc_list = []
                for k, v in cbc_data.items():
                    if v: # Only show non-empty values
                        cbc_list.append({
                            'param': str(k).upper(),
                            'value': v,
                            'range': ranges.get(k, '-')
                        })
                visit.cbc_structured_list = cbc_list

            # 2. Comprehensive Results Handling
            sections = [
                ('serology', SEROLOGY_PARAMS),
                ('biochemistry', BIOCHEMISTRY_PARAMS),
                ('urinalysis', URINALYSIS_PARAMS),
                ('microscopy', MICROSCOPY_PARAMS),
                ('microbiology', MICROBIOLOGY_PARAMS),
            ]
            
            visit.lab_categorized_results = {}
            for sec_key, params in sections:
                sec_data = res.result_data.get(sec_key)
                if isinstance(sec_data, dict) and any(sec_data.values()): # Only add if section has any data
                    current_sec_results = []
                    for p in params:
                        val_raw = sec_data.get(p['key'])
                        if val_raw:
                            val = val_raw.get('value') if isinstance(val_raw, dict) else val_raw
                            remark = val_raw.get('remark') if isinstance(val_raw, dict) else ''
                            if val: # Only show if there's a value
                                current_sec_results.append({
                                    'name': p['name'],
                                    'value': val,
                                    'remark': remark,
                                    'unit': p.get('unit', '-'),
                                    'range': p.get('range', '-')
                                })

                    if current_sec_results:
                        visit.lab_categorized_results[sec_key.title()] = current_sec_results
        # 3. Cancellation Info Handling
        if lab_req and lab_req.status == 'CANCELLED':
            visit.lab_cancel_info = {
                'reason': lab_req.cancel_reason,
                'by': lab_req.cancelled_by.get_full_name() if lab_req.cancelled_by else _('Staff'),
                'at': lab_req.updated_at
            }
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Error enriching lab results: {e}")
        pass

@login_required
@permission_required('medical_records.view_visit', raise_exception=True)
def visit_detail(request, visit_uuid):
    visit = get_object_or_404(
        Visit.objects.visible_to(request.user).select_related(
            'patient', 'doctor', 'checked_in_by', 'triage_nurse', 'current_room'
        ).prefetch_related('logs__performed_by', 'logs__room'),
        uuid=visit_uuid
    )
    enrich_visit_lab_results(visit)
    return render(request, 'medical_records/visit_detail.html', {
        'visit': visit,
    })

@login_required
def visit_detail_ajax(request, visit_uuid):
    visit = get_object_or_404(Visit.objects.visible_to(request.user), uuid=visit_uuid)
    enrich_visit_lab_results(visit)
    return render(request, 'medical_records/visit_detail_modal_content.html', {
        'visit': visit,
    })


@login_required
@permission_required('medical_records.add_visit', raise_exception=True)
def kia_direct_registration(request):
    """
    MCH (Maternal & Child Health) Direct Registration page.
    Supports:
    - GET: Show search + new patient registration form
    - POST (select_patient): Existing patient selected → go to category modal (via redirect with params)
    - POST (register_new): Register brand-new patient → create MCH Visit → examination
    - POST (kia_checkin): Called from patient_list modals for existing patients
    """
    from patients.models import Patient, DailyQueue
    from patients.forms import PatientRegistrationForm
    from django.db.models import Max

    category_labels = {
        'ANA_0_6':     _('MCH - Baby 0-6 Months'),
        'ANA_6_59':    _('MCH - Child 6-59 Months'),
        'IBU_MENYUSU': _('MCH - Breastfeeding Mother'),
        'IBU_HAMIL':   _('MCH - Pregnant Mother'),
    }

    def _get_kia_room():
        return (
            Room.objects.filter(code='KIA').first() or
            Room.objects.filter(name__icontains='KIA').first() or
            Room.objects.filter(name__icontains='Maternidade').first()
        )

    def _create_kia_visit(patient, category, patient_type, weight=None,
                          height=None, muac=None, companion=''):
        kia_room = _get_kia_room()
        if not kia_room:
            return None, _("MCH room not found. Ask admin to create it.")
        queue, _c = DailyQueue.objects.get_or_create(
            date=timezone.localdate(), department='MCH'
        )
        complaint = category_labels.get(category, _('MCH Visit'))
        visit = Visit.objects.create(
            patient=patient,
            current_room=kia_room,
            status='IP',
            queue_number=queue.get_next_number(),
            checked_in_by=request.user,
            doctor=request.user,
            patient_type=patient_type,
            complaint=complaint,
            companion_name=companion,
        )
        
        # Create Vital Signs record for MCH
        VitalSigns.objects.create(
            visit=visit,
            weight=weight or None,
            muac=muac or None,
            kia_category=category,
            height_cm=height or None
        )

        log_visit_action(visit, 'CHECK_IN', request.user, room=kia_room)
        return visit, None

    # ── POST: existing patient from patient_list modals ──────────────
    if request.method == 'POST' and 'patient_uuid' in request.POST and 'kia_category' in request.POST:
        patient_uuid = request.POST.get('patient_uuid')
        category     = request.POST.get('kia_category', '')
        patient_type = request.POST.get('patient_type', 'TUAN')
        patient = get_object_or_404(Patient.objects.visible_to(request.user), uuid=patient_uuid)
        
        # Update patient status based on category
        if category == 'IBU_HAMIL':
            patient.is_pregnant = True
            patient.gender = 'F'
        elif category == 'IBU_MENYUSU':
            patient.is_lactating = True
            patient.gender = 'F'
        elif category.startswith('ANA_'):
            # Children are not pregnant or lactating
            patient.is_pregnant = False
            patient.is_lactating = False
        patient.save()

        visit, err = _create_kia_visit(
            patient, category, patient_type,
            weight=request.POST.get('weight'),
            height=request.POST.get('height'),
            muac=request.POST.get('muac'),
            companion=request.POST.get('companion_name', ''),
        )
        if err:
            messages.error(request, err)
            return redirect('patient_list')
        messages.success(request, _("%(name)s → MCH Queue #%(q)s") % {
            'name': patient.full_name, 'q': visit.queue_number})
        return redirect('perform_examination', visit_uuid=visit.uuid)

    # ── POST: select existing patient from search results ────────────
    if request.method == 'POST' and 'select_patient' in request.POST:
        patient_uuid = request.POST.get('patient_uuid')
        category     = request.POST.get('kia_category', 'ANA_0_6')
        patient_type = request.POST.get('patient_type', 'TUAN')
        patient = get_object_or_404(Patient.objects.visible_to(request.user), uuid=patient_uuid)
        
        # Update patient status based on category
        if category == 'IBU_HAMIL':
            patient.is_pregnant = True
            patient.gender = 'F'
        elif category == 'IBU_MENYUSU':
            patient.is_lactating = True
            patient.gender = 'F'
        elif category.startswith('ANA_'):
            patient.is_pregnant = False
            patient.is_lactating = False
        patient.save()

        visit, err = _create_kia_visit(patient, category, patient_type)
        if err:
            messages.error(request, err)
            return redirect('kia_direct_registration')
        messages.success(request, _("%(name)s → MCH Queue #%(q)s") % {
            'name': patient.full_name, 'q': visit.queue_number})
        return redirect('perform_examination', visit_uuid=visit.uuid)

    # ── POST: register new patient + KIA visit ───────────────────────
    if request.method == 'POST' and 'register_new' in request.POST:
        form = PatientRegistrationForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            year   = timezone.localdate().year
            prefix = f"MD{year}"
            last = Patient.objects.filter(patient_id__startswith=prefix).aggregate(max_id=Max('patient_id'))
            
            # Safe parsing of the last ID number
            try:
                if last['max_id']:
                    new_num = int(last['max_id'][-4:]) + 1
                else:
                    new_num = 1
            except (ValueError, TypeError):
                new_num = 1
                
            patient.patient_id = f"{prefix}{new_num:04d}"
            
            category = request.POST.get('kia_category', 'ANA_0_6')
            if category == 'IBU_HAMIL':
                patient.is_pregnant = True
                patient.gender = 'F'  # Force female for pregnant
            elif category == 'IBU_MENYUSU':
                patient.is_lactating = True
                patient.gender = 'F'  # Force female for breastfeeding
                
            patient.save()

            patient_type = 'FOUN'  # always new
            visit, err = _create_kia_visit(
                patient, category, patient_type,
                weight=request.POST.get('weight'),
                height=request.POST.get('height'),
                muac=request.POST.get('muac'),
                companion=request.POST.get('companion_name', ''),
            )
            if err:
                messages.error(request, err)
                return redirect('kia_direct_registration')
            messages.success(request,
                _("New patient %(name)s registered and admitted to MCH. Queue #%(q)s.") % {
                    'name': patient.full_name, 'q': visit.queue_number})
            return redirect('perform_examination', visit_uuid=visit.uuid)
        # Form invalid — fall through to re-render with errors
        query = request.GET.get('q', '').strip()
        patients = []
        registration_form = form
        return render(request, 'medical_records/kia_direct_registration.html', {
            'registration_form': registration_form,
            'patients': patients,
            'query': query,
            'category_labels': category_labels,
        })

    # ── GET: show page ────────────────────────────────────────────────
    query = request.GET.get('q', '').strip()
    patients = []
    if query:
        patients = Patient.objects.visible_to(request.user).filter(
            Q(full_name__icontains=query) |
            Q(patient_id__icontains=query) |
            Q(phone_number__icontains=query)
        )[:10]

    next_id = Patient.generate_next_id()
    registration_form = PatientRegistrationForm(initial={'patient_id': next_id})
    return render(request, 'medical_records/kia_direct_registration.html', {
        'registration_form': registration_form,
        'patients': patients,
        'query': query,
        'category_labels': category_labels,
    })


# --- Diagnosis Master Data CRUD ---

class DiagnosisListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Diagnosis
    permission_required = 'medical_records.view_diagnosis'
    template_name = 'master_data/list.html'
    context_object_name = 'items'
    paginate_by = 50
    extra_context = {'title': _('Diagnosis List'), 'model_name': 'diagnosis'}

    def get_queryset(self):
        from django.db.models import Q
        queryset = super().get_queryset().select_related('parent', 'category').order_by('code')
        q = self.request.GET.get('q')
        show = self.request.GET.get('show')
        
        if show == 'parents':
            queryset = queryset.filter(parent__isnull=True)
            
        if q:
            queryset = queryset.filter(
                Q(code__icontains=q) | Q(name__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['show'] = self.request.GET.get('show', '')
        return context

@login_required
@permission_required('medical_records.view_diagnosis', raise_exception=True)
def export_diagnoses_pdf(request):
    from clinic_core.pdf_utils import render_to_pdf
    from django.db.models import Q
    from django.utils import timezone
    
    q = request.GET.get('q')
    queryset = Diagnosis.objects.select_related('parent').order_by('code')
    if q:
        queryset = queryset.filter(Q(code__icontains=q) | Q(name__icontains=q))
    
    context = {
        'items': queryset,
        'query': q,
        'today': timezone.now(),
        'clinic_name': _("Clinic Bairo Pite Lanud")
    }
    pdf = render_to_pdf('medical_records/pdf/diagnosis_list_pdf.html', context)
    if pdf:
        filename = f"Diagnosis_List_{timezone.now().strftime('%Y%m%d')}.pdf"
        pdf['Content-Disposition'] = f'attachment; filename="{filename}"'
        return pdf
    return HttpResponse("Error generating PDF", status=500)

@login_required
@permission_required('medical_records.view_diagnosis', raise_exception=True)
def export_diagnoses_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from django.db.models import Q
    from django.http import HttpResponse
    from django.utils import timezone
    
    q = request.GET.get('q')
    queryset = Diagnosis.objects.select_related('parent').order_by('code')
    if q:
        queryset = queryset.filter(Q(code__icontains=q) | Q(name__icontains=q))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diagnoses"
    
    # Header
    headers = ["Diagnosis Code", "Description", "Parent Code"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    # Data
    for row_num, item in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=item.code)
        ws.cell(row=row_num, column=2, value=item.name)
        ws.cell(row=row_num, column=3, value=item.parent.code if item.parent else "")
        
    # Column width
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Diagnosis_List_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

class DiagnosisDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Diagnosis
    permission_required = 'medical_records.view_diagnosis'
    template_name = 'master_data/detail.html'
    context_object_name = 'item'
    extra_context = {'title': _('Diagnosis Detail'), 'model_name': 'diagnosis'}

class DiagnosisCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Diagnosis
    fields = ['code', 'name', 'parent']
    permission_required = 'medical_records.add_diagnosis'
    template_name = 'master_data/form.html'
    success_url = reverse_lazy('diagnosis_list')
    extra_context = {'title': _('Add New Diagnosis')}

class DiagnosisUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Diagnosis
    fields = ['code', 'name', 'parent']
    permission_required = 'medical_records.change_diagnosis'
    template_name = 'master_data/form.html'
    success_url = reverse_lazy('diagnosis_list')
    extra_context = {'title': _('Edit Diagnosis')}

class DiagnosisDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Diagnosis
    permission_required = 'medical_records.delete_diagnosis'
    template_name = 'master_data/confirm_delete.html'
    success_url = reverse_lazy('diagnosis_list')
    extra_context = {'title': _('Delete Diagnosis')}

# --- Room Master Data CRUD ---

class RoomListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Room
    permission_required = 'medical_records.view_room'
    template_name = 'master_data/list.html'
    context_object_name = 'items'
    extra_context = {'title': _('Clinic Rooms'), 'model_name': 'room'}

class RoomDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Room
    permission_required = 'medical_records.view_room'
    template_name = 'master_data/detail.html'
    context_object_name = 'item'
    extra_context = {'title': _('Room Detail'), 'model_name': 'room'}

class RoomCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Room
    fields = ['name', 'code', 'description', 'order']
    permission_required = 'medical_records.add_room'
    template_name = 'master_data/form.html'
    success_url = reverse_lazy('room_list')
    extra_context = {'title': _('Add New Room')}

class RoomUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Room
    fields = ['name', 'code', 'description', 'order']
    permission_required = 'medical_records.change_room'
    template_name = 'master_data/form.html'
    success_url = reverse_lazy('room_list')
    extra_context = {'title': _('Edit Room')}

class RoomDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Room
    permission_required = 'medical_records.delete_room'
    template_name = 'master_data/confirm_delete.html'
    success_url = reverse_lazy('room_list')
    extra_context = {'title': _('Delete Room')}

@login_required
@permission_required('medical_records.view_menu_medical_records', raise_exception=True)
def department_completed_list(request):
    """
    Shows a paginated list of all completed patients per department.
    Dept: emergency, triage, or the specific room code.
    """
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    dept = request.GET.get('dept', '')
    query = request.GET.get('q', '').strip()
    
    # NEW: Default to specialized dept if no dept specified and user belongs to that dept
    staff_profile = getattr(request.user, 'staff_profile', None)
    if not dept and staff_profile and staff_profile.department:
        dept_code = staff_profile.department.code.upper()
        if dept_code in ['HIV', 'AIDS']:
            dept = 'HIV'
        elif dept_code == 'NUTRISI':
            dept = 'NUTRISI'
        elif dept_code == 'TB':
            dept = 'TB'
        elif dept_code == 'KIA':
            dept = 'KIA'
        elif dept_code == 'DENTAL':
            dept = 'DENTAL'
        elif dept_code == 'USG':
            dept = 'USG'
        elif dept_code in ['FAR', 'PHA', 'PHARMACY']:
            dept = 'FAR'
        elif dept_code in ['LAB', 'LABORATORY']:
            dept = 'LAB'
        elif dept_code in ['IGD', 'EMERGENCY']:
            dept = 'emergency'
    
    visits = Visit.objects.visible_to(request.user).select_related('patient', 'current_room').order_by('-visit_date')
    
    page_title = _("All Historical Visits")
    
    if dept == 'triage':
        visits = visits.filter(triage_nurse__isnull=False)
        page_title = _("Triage History")
    elif dept == 'emergency':
        visits = visits.filter(current_room__code__in=['IGD', 'EMERGENCY'], status__in=['COM', 'IP']).exclude(doctor__isnull=True)
        page_title = _("Emergency History")
    elif dept == 'HIV':
        # Master Log: Show ALL historical visits for any patient tagged as HIV
        visits = visits.filter(patient__is_hiv_patient=True, status='COM')
        page_title = _("Complete HIV Patients History")
    elif dept == 'KIA':
        # KIA Master Log: Pregnant/Lactating patients OR visits to KIA room
        visits = visits.filter(
            Q(patient__is_pregnant=True) | Q(patient__is_lactating=True) | Q(current_room__code='KIA'),
            status='COM'
        )
        page_title = _("Complete MCH (KIA) Patients History")
    elif dept == 'NUTRISI':
        # Nutrition Master Log: Show visits for patients who are Pregnant, Lactating, or Children (<5 years)
        from dateutil.relativedelta import relativedelta
        today = timezone.localdate()
        child_limit = today - relativedelta(months=59)
        visits = visits.filter(
            Q(patient__is_pregnant=True) | 
            Q(patient__is_lactating=True) | 
            Q(patient__date_of_birth__gte=child_limit),
            status='COM'
        )
        page_title = _("Nutrition Patients Master History")
    elif dept == 'TB':
        # TB Master Log: Show ALL historical visits for any patient tagged as TB
        visits = visits.filter(patient__is_tb_patient=True, status='COM')
        page_title = _("Complete TB Patients History")
    elif dept == 'LAB':
        visits = visits.filter(current_room__code__in=['LAB', 'LABORATORY'], status='COM')
        page_title = _("Laboratory Visit History")
    elif dept == 'FAR':
        visits = visits.filter(current_room__code__in=['FAR', 'PHA', 'PHARMACY'], status='COM')
        page_title = _("Pharmacy Visit History")
    elif dept:
        # For KIA, Dental, etc (Room-specific history)
        visits = visits.filter(current_room__code=dept, status='COM')
        room_obj = Room.objects.filter(code=dept).first()
        if room_obj:
            page_title = f"{room_obj.name} - {_('History')}"
        else:
            page_title = f"{dept} - {_('History')}"
    else:
        # General completed
        visits = visits.filter(status='COM')
        
    if query:
        visits = visits.filter(
            Q(patient__full_name__icontains=query) | 
            Q(patient__patient_id__icontains=query)
        )
        
    paginator = Paginator(visits, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'medical_records/completed_list.html', {
        'page_obj': page_obj,
        'page_title': page_title,
        'dept': dept,
        'query': query,
    })

@login_required
@permission_required('medical_records.view_visit', raise_exception=True)
def export_visit_history_pdf(request):
    from django.db.models import Q
    from clinic_core.pdf_utils import render_to_pdf
    
    dept = request.GET.get('dept', '')
    query = request.GET.get('q', '').strip()
    visits = Visit.objects.visible_to(request.user).select_related('patient', 'current_room', 'doctor').order_by('-visit_date')
    
    page_title = _("All Visits")
    if dept == 'triage':
        visits = visits.filter(triage_nurse__isnull=False)
        page_title = _("Triage History")
    elif dept == 'HIV':
        visits = visits.filter(patient__is_hiv_patient=True, status='COM')
        page_title = _("HIV Patients History")
    elif dept == 'NUTRISI':
        from dateutil.relativedelta import relativedelta
        today = timezone.localdate()
        child_limit = today - relativedelta(months=59)
        visits = visits.filter(
            Q(patient__is_pregnant=True) | 
            Q(patient__is_lactating=True) | 
            Q(patient__date_of_birth__gte=child_limit),
            status='COM'
        )
        page_title = _("Nutrition Patients History")
    elif dept:
        visits = visits.filter(current_room__code=dept)
        room_obj = Room.objects.filter(code=dept).first()
        page_title = f"{room_obj.name} - {_('History')}" if room_obj else f"{dept} - {_('History')}"

    if query:
        visits = visits.filter(
            Q(patient__full_name__icontains=query) | 
            Q(patient__patient_id__icontains=query)
        )
    
    # We limit to 500 for PDF to avoid timeout, but usually history is paged
    visits = visits[:500]
    
    context = {
        'visits': visits,
        'page_title': page_title,
        'query': query,
    }
    pdf = render_to_pdf('medical_records/pdf/visit_history_pdf.html', context)
    if pdf:
        filename = f"Visit_History_{timezone.now().strftime('%Y%m%d')}.pdf"
        pdf['Content-Disposition'] = f'attachment; filename="{filename}"'
        return pdf
    return HttpResponse("Error generating PDF", status=500)

@login_required
@permission_required('medical_records.view_visit', raise_exception=True)
def export_visit_history_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.db.models import Q

    dept = request.GET.get('dept', '')
    query = request.GET.get('q', '').strip()
    visits = Visit.objects.visible_to(request.user).select_related('patient', 'current_room', 'doctor').order_by('-visit_date')

    if dept == 'triage':
        visits = visits.filter(triage_nurse__isnull=False)
    elif dept == 'emergency':
        visits = visits.filter(current_room__code__in=['IGD', 'EMERGENCY'], status__in=['COM', 'IP']).exclude(doctor__isnull=True)
    elif dept == 'NUTRISI':
        from dateutil.relativedelta import relativedelta
        today = timezone.localdate()
        child_limit = today - relativedelta(months=59)
        visits = visits.filter(
            Q(patient__is_pregnant=True) | 
            Q(patient__is_lactating=True) | 
            Q(patient__date_of_birth__gte=child_limit),
            status='COM'
        )
    elif dept:
        visits = visits.filter(current_room__code=dept)

    if query:
        visits = visits.filter(
            Q(patient__full_name__icontains=query) | 
            Q(patient__patient_id__icontains=query)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Visit History"

    # Header row
    headers = [
        _('Date'), _('Patient ID'), _('Patient Name'), _('Department'), _('Doctor'), _('Diagnosis'), _('Status')
    ]
    
    header_fill = PatternFill(start_color='5D2D91', end_color='5D2D91', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=str(header))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_num, visit in enumerate(visits[:1000], 2):
        ws.cell(row=row_num, column=1, value=visit.visit_date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_num, column=2, value=visit.patient.patient_id)
        ws.cell(row=row_num, column=3, value=visit.patient.full_name)
        ws.cell(row=row_num, column=4, value=visit.current_room.name if visit.current_room else "-")
        ws.cell(row=row_num, column=5, value=visit.doctor.get_full_name() if visit.doctor else "-")
        # Handle multiple diagnoses for Excel
        primary_diag = visit.visit_diagnoses.filter(is_primary=True).first()
        if not primary_diag:
            primary_diag = visit.visit_diagnoses.first()
            
        diag_str = f"{primary_diag.diagnosis.code} - {primary_diag.diagnosis.name}" if primary_diag else "-"
        ws.cell(row=row_num, column=6, value=diag_str)
        ws.cell(row=row_num, column=7, value=str(visit.get_status_display()))

    # Column width
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Visit_History_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@permission_required('medical_records.view_visit', raise_exception=True)
def visit_summary_pdf(request, visit_uuid):
    """Generate PDF summary for a single visit."""
    visit = get_object_or_404(
        Visit.objects.visible_to(request.user).select_related(
            'patient', 'doctor', 'checked_in_by', 'triage_nurse', 'current_room', 'diagnosis'
        ).prefetch_related(
            'logs__performed_by', 
            'logs__room',
            'secondary_diagnoses'
        ),
        uuid=visit_uuid
    )
    enrich_visit_lab_results(visit)
    from clinic_core.pdf_utils import render_to_pdf
    context = {'visit': visit}
    pdf = render_to_pdf('medical_records/pdf/visit_summary_pdf.html', context)
    if pdf:
        pdf['Content-Disposition'] = f'inline; filename="Visit_{visit.patient.patient_id}_{visit.visit_date.date()}.pdf"'
        return pdf
    from django.http import HttpResponse
    return HttpResponse("Error generating PDF", status=500)

@login_required
def search_diagnosis_ajax(request):
    """AJAX endpoint for Select2 diagnosis search."""
    from django.http import JsonResponse
    from django.db.models import Q
    
    query = request.GET.get('q', '')
    try:
        page = int(request.GET.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    page_size = 30
    
    start = (page - 1) * page_size
    end = start + page_size
    
    queryset = Diagnosis.objects.all()
    if query:
        queryset = queryset.filter(
            Q(code__icontains=query) | Q(name__icontains=query)
        )
    
    total_count = queryset.count()
    results = queryset.order_by('code')[start:end]
    
    data = {
        'results': [{'id': d.id, 'text': f"{d.code} - {d.name}"} for d in results],
        'total_count': total_count
    }
    return JsonResponse(data)


@login_required
@permission_required('medical_records.view_menu_specialist_nutrition', raise_exception=True)
def nutrition_statistics(request):
    """
    Halaman statistika nutrisi format TLHIS/04 untuk laporan ke Ministerio da Saúde.
    Data yang tersedia: MUAC, registrasi pasien, pasien baru.
    Data yang belum ada (—): tinggi badan, Z-score, suplemen, PTE/PTI.
    """
    from dateutil.relativedelta import relativedelta
    from django.db.models import Count, Q
    from patients.models import Patient

    # --- Filter Bulan ---
    selected_month = request.GET.get('month', timezone.localdate().strftime('%Y-%m'))
    try:
        year, month = int(selected_month[:4]), int(selected_month[5:7])
    except (ValueError, IndexError):
        year, month = timezone.localdate().year, timezone.localdate().month

    month_start = timezone.datetime(year, month, 1).date()
    if month == 12:
        month_end = timezone.datetime(year + 1, 1, 1).date()
    else:
        month_end = timezone.datetime(year, month + 1, 1).date()

    # --- Batas usia ---
    # NEW: Use month_end as reference so historical reports are consistent
    reference_date = month_end 
    
    limit_0m  = reference_date - relativedelta(months=0)
    limit_24m = reference_date - relativedelta(months=24)
    limit_59m = reference_date - relativedelta(months=59)

    # Pasien anak berdasarkan kelompok usia
    def children_qs(age_min_m, age_max_m, gender=None):
        dob_upper = reference_date - relativedelta(months=age_min_m)
        dob_lower = reference_date - relativedelta(months=age_max_m)
        qs = Patient.objects.filter(
            date_of_birth__lte=dob_upper,
            date_of_birth__gt=dob_lower,
        )
        if gender:
            qs = qs.filter(gender=gender)
        return qs

    # Kunjungan NUTRISI pada bulan ini
    nutri_visits_month = Visit.objects.filter(
        current_room__code='NUTRISI',
        visit_date__date__gte=month_start,
        visit_date__date__lt=month_end,
    )

    def count_visits(age_min_m, age_max_m, gender, qs=None):
        if qs is None:
            qs = nutri_visits_month
        dob_upper = reference_date - relativedelta(months=age_min_m)
        dob_lower = reference_date - relativedelta(months=age_max_m)
        return qs.filter(
            patient__date_of_birth__lte=dob_upper,
            patient__date_of_birth__gt=dob_lower,
            patient__gender=gender,
        ).values('patient').distinct().count()

    def count_new_visits(age_min_m, age_max_m, gender):
        dob_upper = reference_date - relativedelta(months=age_min_m)
        dob_lower = reference_date - relativedelta(months=age_max_m)
        return nutri_visits_month.filter(
            patient__date_of_birth__lte=dob_upper,
            patient__date_of_birth__gt=dob_lower,
            patient__gender=gender,
            patient_type='FOUN',
        ).values('patient').distinct().count()

    # MUAC classification — only for visits with MUAC recorded
    def count_muac(gender, muac_min=None, muac_max=None):
        qs = nutri_visits_month.filter(
            patient__date_of_birth__lte=reference_date - relativedelta(months=6),
            patient__date_of_birth__gt=reference_date - relativedelta(months=59),
            patient__gender=gender,
            muac__isnull=False,
        )
        if muac_min is not None:
            qs = qs.filter(muac__gte=muac_min)
        if muac_max is not None:
            qs = qs.filter(muac__lt=muac_max)
        return qs.count()

    s = {
        # --- Registrasi (1-3) ---
        'reg_0_23_m':  count_visits(0, 23, 'M'),
        'reg_0_23_f':  count_visits(0, 23, 'F'),
        'reg_24_59_m': count_visits(24, 59, 'M'),
        'reg_24_59_f': count_visits(24, 59, 'F'),

        'new_0_23_m':  count_new_visits(0, 23, 'M'),
        'new_0_23_f':  count_new_visits(0, 23, 'F'),
        'new_24_59_m': count_new_visits(24, 59, 'M'),
        'new_24_59_f': count_new_visits(24, 59, 'F'),

        'mon_0_23_m':  count_visits(0, 23, 'M'),
        'mon_0_23_f':  count_visits(0, 23, 'F'),
        'mon_24_59_m': count_visits(24, 59, 'M'),
        'mon_24_59_f': count_visits(24, 59, 'F'),

        # --- WAZ / WHZ / HAZ — belum ada data (None = tampil "—") ---
        'waz_norm_0_23_m': None, 'waz_norm_0_23_f': None,
        'waz_norm_24_59_m': None, 'waz_norm_24_59_f': None,
        'waz_mod_0_23_m': None, 'waz_mod_0_23_f': None,
        'waz_mod_24_59_m': None, 'waz_mod_24_59_f': None,
        'waz_sev_0_23_m': None, 'waz_sev_0_23_f': None,
        'waz_sev_24_59_m': None, 'waz_sev_24_59_f': None,

        'whz_norm_0_23_m': None, 'whz_norm_0_23_f': None,
        'whz_norm_24_59_m': None, 'whz_norm_24_59_f': None,
        'whz_mod_0_23_m': None, 'whz_mod_0_23_f': None,
        'whz_mod_24_59_m': None, 'whz_mod_24_59_f': None,
        'whz_sev_0_23_m': None, 'whz_sev_0_23_f': None,
        'whz_sev_24_59_m': None, 'whz_sev_24_59_f': None,
        'whz_ow_0_23_m': None, 'whz_ow_0_23_f': None,
        'whz_ow_24_59_m': None, 'whz_ow_24_59_f': None,
        'whz_ob_0_23_m': None, 'whz_ob_0_23_f': None,
        'whz_ob_24_59_m': None, 'whz_ob_24_59_f': None,

        'haz_norm_0_23_m': None, 'haz_norm_0_23_f': None,
        'haz_norm_24_59_m': None, 'haz_norm_24_59_f': None,
        'haz_mod_0_23_m': None, 'haz_mod_0_23_f': None,
        'haz_mod_24_59_m': None, 'haz_mod_24_59_f': None,
        'haz_sev_0_23_m': None, 'haz_sev_0_23_f': None,
        'haz_sev_24_59_m': None, 'haz_sev_24_59_f': None,

        # --- Konseling (7) ---
        'counsel_0_23_m': None, 'counsel_0_23_f': None,
        'counsel_24_59_m': None, 'counsel_24_59_f': None,

        # --- MUAC (8) — DATA TERSEDIA ---
        'muac_norm_m': count_muac('M', muac_min=12.5),
        'muac_norm_f': count_muac('F', muac_min=12.5),
        'muac_mam_m':  count_muac('M', muac_min=11.5, muac_max=12.5),
        'muac_mam_f':  count_muac('F', muac_min=11.5, muac_max=12.5),
        'muac_sam_m':  count_muac('M', muac_max=11.5),
        'muac_sam_f':  count_muac('F', muac_max=11.5),

        # --- PTE / PTI / Suplemen — belum ada ---
        'pte_cured_m': None, 'pte_cured_f': None,
        'pte_not_cured_m': None, 'pte_not_cured_f': None,
        'pte_default_m': None, 'pte_default_f': None,
        'pte_dead_m': None, 'pte_dead_f': None,
        'pti_m': None, 'pti_f': None,
        'vita_6_11_m': None, 'vita_6_11_f': None,
        'vita_12_59_m': None, 'vita_12_59_f': None,
        'alumb_12_23_m': None, 'alumb_12_23_f': None,
        'alumb_24_59_m': None, 'alumb_24_59_f': None,
        'mnr_6_23_m': None, 'mnr_6_23_f': None,
    }

    return render(request, 'medical_records/nutrition_statistics.html', {
        'selected_month': selected_month,
        's': s,
        'missing_fields': True,
    })

@login_required
@permission_required('medical_records.view_menu_specialist_tb', raise_exception=True)
def tb_screening_list(request):
    from .models import TBScreening
    from patients.models import Municipio, PostoAdministrativo, Suco, Aldeia
    
    screenings = TBScreening.objects.all().select_related('patient', 'municipio', 'posto', 'suco', 'aldeia', 'screened_by')
    
    # Text Search
    q = request.GET.get('q', '')
    if q:
        screenings = screenings.filter(
            Q(full_name__icontains=q) | 
            Q(patient__full_name__icontains=q) |
            Q(outreach_location__icontains=q) |
            Q(phone_number__icontains=q)
        )
    
    # Geography Filters
    m_id = request.GET.get('municipio')
    p_id = request.GET.get('posto')
    s_id = request.GET.get('suco')
    a_id = request.GET.get('aldeia')
    year = request.GET.get('year')
    
    if m_id: screenings = screenings.filter(municipio_id=m_id)
    if p_id: screenings = screenings.filter(posto_id=p_id)
    if s_id: screenings = screenings.filter(suco_id=s_id)
    if a_id: screenings = screenings.filter(aldeia_id=a_id)
    if year: screenings = screenings.filter(screening_date__year=year)

    # Calculate stats for the dashboard (after filtering)
    suspect_count = screenings.filter(is_suspect=True).count()
    sputum_count = screenings.filter(sputum_collected=True).count()
    referred_count = screenings.exclude(referral_status='NONE').count()
    positive_count = screenings.filter(lab_result='POSITIVE').count()
    
    # Context data for filters
    municipios = Municipio.objects.all()
    # Available years
    years = TBScreening.objects.dates('screening_date', 'year', order='DESC')
    
    return render(request, 'medical_records/tb_screening_list.html', {
        'screenings': screenings,
        'q': q,
        'suspect_count': suspect_count,
        'sputum_count': sputum_count,
        'referred_count': referred_count,
        'positive_count': positive_count,
        'municipios': municipios,
        'years': [y.year for y in years],
        'selected_filters': {
            'municipio': m_id, 'posto': p_id, 'suco': s_id, 'aldeia': a_id, 'year': year
        }
    })

@login_required
@permission_required('medical_records.view_menu_specialist_tb', raise_exception=True)
def tb_screening_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from .models import TBScreening
    
    # Reuse filter logic from list view (simplified for now)
    screenings = TBScreening.objects.all().select_related('patient', 'municipio', 'posto', 'suco', 'aldeia', 'screened_by')
    # (Apply same filters as above if needed, or export current queryset)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TB Outreach Screening"
    
    headers = [
        'Date', 'Patient Name', 'Phone', 'Gender', 'Age', 
        'Municipio', 'Posto', 'Suco', 'Aldeia', 'Site',
        'Suspect?', 'Symptoms', 'Sputum Taken', 'Lab Result', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row, s in enumerate(screenings, 2):
        name = s.patient.full_name if s.patient else s.full_name
        symptoms = []
        if s.has_cough_2_weeks: symptoms.append("Cough")
        if s.has_fever: symptoms.append("Fever")
        if s.has_night_sweats: symptoms.append("Night Sweats")
        if s.has_weight_loss: symptoms.append("Weight Loss")
        
        ws.cell(row=row, column=1, value=s.screening_date)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=s.phone_number or (s.patient.phone_number if s.patient else ""))
        ws.cell(row=row, column=4, value=s.gender or (s.patient.gender if s.patient else ""))
        ws.cell(row=row, column=5, value=s.age or (s.patient.age if s.patient else ""))
        ws.cell(row=row, column=6, value=str(s.municipio or ""))
        ws.cell(row=row, column=7, value=str(s.posto or ""))
        ws.cell(row=row, column=8, value=str(s.suco or ""))
        ws.cell(row=row, column=9, value=str(s.aldeia or ""))
        ws.cell(row=row, column=10, value=s.outreach_location)
        ws.cell(row=row, column=11, value="YES" if s.is_suspect else "No")
        ws.cell(row=row, column=12, value=", ".join(symptoms))
        ws.cell(row=row, column=13, value="YES" if s.sputum_collected else "No")
        ws.cell(row=row, column=14, value=s.get_lab_result_display())
        ws.cell(row=row, column=15, value=s.notes)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="TB_Screening_Report_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response

@login_required
@permission_required('medical_records.view_menu_specialist_tb', raise_exception=True)
def tb_screening_export_pdf(request):
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from .models import TBScreening
    
    screenings = TBScreening.objects.all().select_related('patient', 'municipio', 'posto', 'suco', 'aldeia')
    
    html_string = render_to_string('medical_records/tb_screening_pdf.html', {
        'screenings': screenings,
        'report_date': timezone.now(),
    })
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="TB_Screening_Report_{timezone.now().date()}.pdf"'
    
    # Create a PDF
    pisa_status = pisa.CreatePDF(html_string, dest=response)
    
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html_string + '</pre>')
    return response

@login_required
@permission_required('medical_records.view_menu_specialist_tb', raise_exception=True)
def tb_screening_create(request):
    from .forms import TBScreeningForm
    if request.method == 'POST':
        form = TBScreeningForm(request.POST)
        if form.is_valid():
            screening = form.save(commit=False)
            screening.screened_by = request.user
            screening.save()
            messages.success(request, _("TB Screening record saved successfully."))
            return redirect('tb_screening_list')
    else:
        form = TBScreeningForm(initial={'screening_date': timezone.localdate()})
    
    return render(request, 'medical_records/tb_screening_form.html', {
        'form': form,
        'title': _("New TB Outreach Screening")
    })

@login_required
@permission_required('medical_records.view_menu_specialist_tb', raise_exception=True)
def tb_screening_edit(request, uuid):
    from .models import TBScreening
    from .forms import TBScreeningForm
    screening = get_object_or_404(TBScreening, uuid=uuid)
    
    if request.method == 'POST':
        form = TBScreeningForm(request.POST, instance=screening)
        if form.is_valid():
            form.save()
            messages.success(request, _("TB Screening record updated successfully."))
            return redirect('tb_screening_list')
    else:
        form = TBScreeningForm(instance=screening)
    
    return render(request, 'medical_records/tb_screening_form.html', {
        'form': form,
        'title': _("Edit TB Outreach Screening"),
        'is_edit': True
    })

# =============================================
# TB TREATMENT MANAGEMENT (DOTS)
# =============================================

@login_required
def tb_treatment_dashboard(request):
    from .models import TBCase, Visit
    from django.db.models import Count, Q

    today = timezone.localdate()

    cases = TBCase.objects.all().select_related('patient').order_by('-date_started')

    # Simple search
    q = request.GET.get('q', '')
    if q:
        cases = cases.filter(
            Q(patient__full_name__icontains=q) |
            Q(tb_registration_number__icontains=q)
        )

    # --- Visit statistics for TB room (Optimized with Aggregation) ---
    from django.db.models import Count, Q
    tb_visits_qs = Visit.objects.filter(current_room__code='TB')
    
    case_stats = TBCase.objects.aggregate(
        active=Count('uuid', filter=Q(is_active=True)),
        inactive=Count('uuid', filter=Q(is_active=False))
    )
    
    visit_stats = tb_visits_qs.aggregate(
        today=Count('uuid', filter=Q(visit_date__date=today)),
        month=Count('uuid', filter=Q(visit_date__year=today.year, visit_date__month=today.month)),
        total=Count('uuid')
    )

    stats = {
        'active_cases':    case_stats['active'],
        'inactive_cases':  case_stats['inactive'],
        'visits_today':    visit_stats['today'],
        'visits_month':    visit_stats['month'],
        'visits_total':    visit_stats['total'],
    }

    # --- Pagination (MOVED UP for performance) ---
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(cases, 15) # Show 15 cases per page
    page = request.GET.get('page')
    try:
        cases_page = paginator.page(page)
    except PageNotAnInteger:
        cases_page = paginator.page(1)
    except EmptyPage:
        cases_page = paginator.page(paginator.num_pages)

    # --- Visit Count Calculation (Only for the current page) ---
    # This is much faster because it only processes 15 patients instead of all.
    page_patient_ids = [c.patient_id for c in cases_page]
    visit_counts = (
        Visit.objects.filter(patient_id__in=page_patient_ids, current_room__code='TB')
        .values('patient_id')
        .annotate(cnt=Count('uuid'))
    )
    visit_map = {row['patient_id']: row['cnt'] for row in visit_counts}
    for case in cases_page:
        case.tb_visit_count = visit_map.get(case.patient_id, 0)

    return render(request, 'medical_records/tb_treatment_dashboard.html', {
        'cases': cases_page,
        'query': q,
        'stats': stats,
        'today': today,
    })

@login_required
def tb_case_create(request):
    from .forms import TBCaseForm
    if request.method == 'POST':
        form = TBCaseForm(request.POST)
        if form.is_valid():
            case = form.save()
            
            # Auto-check-in to TB Room for today if not already there
            from .models import Visit, Room
            from patients.models import DailyQueue
            today = timezone.localdate()
            tb_room = Room.objects.filter(code='TB').first()
            
            # Check if patient already has a visit today in TB room
            existing_visit = Visit.objects.filter(patient=case.patient, visit_date__date=today, current_room=tb_room).exists()
            if not existing_visit and tb_room:
                queue, created = DailyQueue.objects.get_or_create(date=today, department='TB')
                Visit.objects.create(
                    patient=case.patient,
                    current_room=tb_room,
                    queue_number=queue.get_next_number(),
                    status='IP', # Set directly to In Progress
                    checked_in_by=request.user,
                )
            
            messages.success(request, _("TB Patient registered to DOTS program successfully."))
            return redirect('tb_treatment_card', case_uuid=case.uuid)
    else:
        # Pre-select patient and pull clinical data from latest visit
        patient_id = request.GET.get('patient')
        initial = {'patient': patient_id} if patient_id else {}
        
        if patient_id:
            from .models import VisitDiagnosis
            from laboratory.models import LabResult
            from radiology.models import RadiologyResult
            
            # Get latest primary diagnosis for this patient
            latest_diag = VisitDiagnosis.objects.filter(
                visit__patient_id=patient_id, 
                is_primary=True
            ).select_related('diagnosis').order_by('-visit__visit_date').first()
            
            if latest_diag:
                code = latest_diag.diagnosis.code.upper()
                # A15-A16: Pulmonary
                if code.startswith('A15') or code.startswith('A16'):
                    initial['classification'] = 'P'
                # A17-A19: Extra-Pulmonary
                elif code.startswith('A17') or code.startswith('A18') or code.startswith('A19'):
                    initial['classification'] = 'EP'
                    initial['site_of_eptb'] = latest_diag.diagnosis.name
                    
            # Pull Lab Data
            latest_lab = LabResult.objects.filter(
                lab_request__visit__patient_id=patient_id,
                lab_request__status='COMPLETED'
            ).order_by('-completed_at').first()
            
            if latest_lab and latest_lab.result_data:
                res = latest_lab.result_data
                hiv = res.get('serology', {}).get('hiv', '')
                if hiv:
                    initial['hiv_status'] = hiv[:50]
                        
                sputum = res.get('microbiology', {}).get('zn_afb', '')
                if sputum:
                    initial['initial_sputum'] = sputum[:50]
                    
                glucose = res.get('biochemistry', {}).get('glucose', {})
                if isinstance(glucose, dict) and glucose.get('value'):
                    initial['diabetes_status'] = glucose.get('value')
                elif isinstance(glucose, str) and glucose:
                    initial['diabetes_status'] = glucose[:50]
            
            # Pull Xray
            latest_rad = RadiologyResult.objects.filter(
                radiology_request__visit__patient_id=patient_id,
                radiology_request__status='COMPLETED'
            ).order_by('-completed_at').first()
            if latest_rad and latest_rad.impression:
                initial['initial_xray'] = latest_rad.impression[:255]
        
        form = TBCaseForm(initial=initial)
        
    from .models import TBCase
    last_case = TBCase.objects.order_by('-created_at').first()
    last_tb_number = last_case.tb_registration_number if last_case else None
        
    return render(request, 'medical_records/tb_case_form.html', {
        'form': form,
        'title': _("Register New TB Case (DOTS)"),
        'last_tb_number': last_tb_number
    })

@login_required
def tb_treatment_card(request, case_uuid):
    from .models import TBCase, TBDailyDose
    from laboratory.models import LabResult
    from radiology.models import RadiologyResult
    
    case = get_object_or_404(TBCase, uuid=case_uuid)
    # Fetch all daily doses for this case
    doses = case.daily_doses.all().order_by('date')
    
    # --- AUTO-PULL DATA FROM OTHER DEPARTMENTS ---
    pulled_data = {
        'clinical_notes': '',
        'hiv_result': '',
        'hiv_date': '',
        'glucose_value': '',
        'afb_result': '',
        'lab_no': '',
        'lab_date': '',
        'xray_date': '',
        'xray_result': '',
        'xray_abnormal': False,
        'visit_diag_code': '',
        'visit_diag_name': '',
        'visit_diag_date': '',
        'suggested_classification': '',
        'suggested_site': '',
    }
    
    # 1. Latest Lab Results (HIV, Diabetes, Sputum/AFB)
    latest_lab = LabResult.objects.filter(
        lab_request__visit__patient=case.patient,
        lab_request__status='COMPLETED'
    ).select_related('lab_request').order_by('-completed_at').first()
    
    if latest_lab and latest_lab.result_data:
        res = latest_lab.result_data
        # HIV
        pulled_data['hiv_result'] = res.get('serology', {}).get('hiv', '')
        pulled_data['hiv_date'] = latest_lab.completed_at
        
        # Diabetes (Check both FBS and RBS logic if needed)
        glu = res.get('biochemistry', {}).get('glucose', {})
        if isinstance(glu, dict):
            pulled_data['glucose_value'] = glu.get('value', '')
        else:
            pulled_data['glucose_value'] = glu
            
        # Sputum/AFB (ZN Stain)
        pulled_data['afb_result'] = res.get('microbiology', {}).get('zn_afb', '')
        pulled_data['lab_no'] = latest_lab.lab_request.lab_no
        pulled_data['lab_date'] = latest_lab.completed_at
        
    # 2. Latest Radiology (Thorax X-Ray)
    latest_rad = RadiologyResult.objects.filter(
        radiology_request__visit__patient=case.patient,
        radiology_request__status='COMPLETED'
    ).select_related('radiology_request').order_by('-completed_at').first()
    
    if latest_rad:
        pulled_data['xray_date'] = latest_rad.completed_at
        pulled_data['xray_result'] = latest_rad.impression
        # If findings indicate abnormality
        pulled_data['xray_abnormal'] = 'abnormal' in (latest_rad.impression or '').lower() or 'positif' in (latest_rad.impression or '').lower()
    
    # 3. Latest Clinical Notes (Discussion)
    from .models import Visit, VisitDiagnosis
    latest_visit = Visit.objects.filter(
        patient=case.patient,
        status='COM'
    ).exclude(clinical_notes__isnull=True).exclude(clinical_notes='').order_by('-visit_date').first()
    
    if latest_visit:
        pulled_data['clinical_notes'] = latest_visit.clinical_notes

    # 4. Pull TB Classification from OPD/IGD VisitDiagnosis (ICD-10 A15–A19)
    # Only suggest if card_data doesn't already have an explicit classification saved
    card_has_classification = case.card_data.get('cl_pul') or case.card_data.get('cl_ep')
    if not card_has_classification:
        tb_diag = VisitDiagnosis.objects.filter(
            visit__patient=case.patient,
            visit__status='COM',
            diagnosis__code__regex=r'^A1[5-9]'
        ).select_related('diagnosis', 'visit').order_by('-visit__visit_date').first()

        if tb_diag:
            code = tb_diag.diagnosis.code.upper()
            pulled_data['visit_diag_code'] = code
            pulled_data['visit_diag_name'] = tb_diag.diagnosis.name
            pulled_data['visit_diag_date'] = tb_diag.visit.visit_date

            # A15, A16 → Pulmonary TB
            if code.startswith('A15') or code.startswith('A16'):
                pulled_data['suggested_classification'] = 'P'
                pulled_data['suggested_site'] = ''
            # A17, A18, A19 → Extra-Pulmonary TB
            elif code.startswith('A17') or code.startswith('A18') or code.startswith('A19'):
                pulled_data['suggested_classification'] = 'EP'
                pulled_data['suggested_site'] = tb_diag.diagnosis.name

    days_list = [str(i).zfill(2) for i in range(1, 32)]
    
    # 4. Pull available TB medicines for dropdowns
    from pharmacy.models import Medicine
    tb_medicines = Medicine.objects.filter(department_category='TB', is_active=True).order_by('name')

    # 5. Count TB visits for this patient
    tb_visit_count = Visit.objects.filter(patient=case.patient, current_room__code='TB').count()
    
    # 6. Count Treatment Logs (Monitoring entries)
    log_count = case.logs.count()

    return render(request, 'medical_records/tb_treatment_card.html', {
        'case': case,
        'doses': doses,
        'today': timezone.localdate(),
        'days_list': days_list,
        'pulled_data': pulled_data,
        'tb_medicines': tb_medicines,
        'tb_visit_count': tb_visit_count,
        'log_count': log_count
    })

@login_required
@permission_required('medical_records.change_tbcase', raise_exception=True)
def tb_card_save(request, case_uuid):
    """
    Saves the entire TB Treatment Card data, including the daily dose grid 
    and unstructured card_data.
    """
    from .models import TBCase, TBDailyDose
    import json
    
    case = get_object_or_404(TBCase, uuid=case_uuid)
    
    if request.method == 'POST':
        # 1. Save Unstructured Card Data (JSON)
        card_data = {}
        for key, value in request.POST.items():
            if key.startswith('card_'):
                clean_key = key.replace('card_', '')
                card_data[clean_key] = value
                
                # 2. Also try to sync with TBDailyDose if it's a dose mark
                # Format: dose_p1_r1_d01
                if clean_key.startswith('dose_'):
                    try:
                        # We try to find the month for this row to derive a real date
                        # Example key: dose_p1_r1_d01 -> row is p1_r1
                        parts = clean_key.split('_') # ['dose', 'p1', 'r1', 'd01']
                        row_prefix = f"month_{parts[1]}_{parts[2]}" # month_p1_r1
                        month_val = request.POST.get(f"card_{row_prefix}", "") # e.g. "2026-05"
                        day_val = parts[3].replace('d', '') # "01"
                        
                        if month_val and len(month_val) >= 7: # YYYY-MM
                            # Only sync if we have a valid-looking month/year
                            date_str = f"{month_val[:7]}-{day_val.zfill(2)}"
                            
                            status_map = {'✔': 'DONE', '-': 'UNOBSERVED', '0': 'MISSED'}
                            db_status = status_map.get(value)
                            
                            if db_status:
                                TBDailyDose.objects.update_or_create(
                                    tb_case=case,
                                    date=date_str,
                                    defaults={'status': db_status, 'recorded_by': request.user}
                                )
                    except Exception:
                        pass # Continue if date parsing fails
        
        # 3. Sync model fields from card_data
        if card_data.get('cl_pul'): case.classification = 'P'
        elif card_data.get('cl_ep'): case.classification = 'EP'
        
        if card_data.get('cl_site'): case.site_of_eptb = card_data.get('cl_site')
        
        # Patient Type mapping
        if card_data.get('tp_foun'): case.patient_type = 'FOUN'
        elif card_data.get('tp_relap'): case.patient_type = 'RELAPSU'
        elif card_data.get('tp_lakon'): case.patient_type = 'DEPOIS_LAKON'
        elif card_data.get('tp_falla'): case.patient_type = 'FALHA'
        elif card_data.get('tp_seluk'): case.patient_type = 'OUTRU'

        # Outcome mapping and auto-deactivate
        if card_data.get('res_kura'): 
            case.outcome = 'CURA'
            case.is_active = False
        elif card_data.get('res_kom'): 
            case.outcome = 'KOMPLETU'
            case.is_active = False
        elif card_data.get('res_mate'): 
            case.outcome = 'MATE'
            case.is_active = False
        elif card_data.get('res_falla'): 
            case.outcome = 'FALHA'
            case.is_active = False
        elif card_data.get('res_lakon'): 
            case.outcome = 'LAKON'
            case.is_active = False
        elif card_data.get('res_trans'): 
            case.outcome = 'TRANSFER'
            case.is_active = False

        # Sync Monitoring Data to model fields (for reporting)
        if card_data.get('s_m0_r'): case.initial_sputum = card_data.get('s_m0_r')
        if card_data.get('s_m2_r'): case.sputum_month_2 = card_data.get('s_m2_r')
        if card_data.get('s_m5_r'): case.sputum_month_5 = card_data.get('s_m5_r')
        if card_data.get('s_m6_r'): case.sputum_month_6 = card_data.get('s_m6_r')
        
        if card_data.get('x_m0_r'): case.initial_xray = card_data.get('x_m0_r')
        
        if card_data.get('w_m0'):
            try:
                from decimal import Decimal
                case.initial_weight = Decimal(card_data.get('w_m0'))
            except:
                pass
        
        case.card_data = card_data
        case.save()

        # Mark any active TB visits from OPD/IGD as Completed
        from .models import Visit
        active_visits = Visit.objects.filter(
            patient=case.patient,
            status__in=['SCH', 'IP'],
            current_room__code='TB'
        )
        for v in active_visits:
            v.status = 'COM'
            v.save()

        messages.success(request, _("TB Treatment Card updated successfully."))
        return redirect('tb_treatment_dashboard')

    return redirect('tb_treatment_card', case_uuid=case.uuid)

@login_required
def tb_treatment_log_add(request, case_uuid):
    """
    Handles adding a daily dose log from the modal in TB Treatment Card.
    """
    from .models import TBCase, TBDailyDose
    case = get_object_or_404(TBCase, uuid=case_uuid)
    
    if request.method == 'POST':
        date_val = request.POST.get('date')
        is_observed = request.POST.get('is_observed') == 'true'
        notes = request.POST.get('notes', '')
        
        # Get or create to avoid duplicates for the same day
        dose, created = TBDailyDose.objects.update_or_create(
            tb_case=case,
            date=date_val,
            defaults={
                'is_observed': is_observed,
                'notes': notes,
                'recorded_by': request.user
            }
        )
        
        if created:
            messages.success(request, _("Daily dose for %s recorded.") % date_val)
        else:
            messages.success(request, _("Daily dose for %s updated.") % date_val)
            
    return redirect('tb_treatment_card', case_uuid=case.uuid)

def get_quarter_dates(year, quarter):
    import datetime
    if quarter == 1:
        return datetime.date(year, 1, 1), datetime.date(year, 3, 31)
    elif quarter == 2:
        return datetime.date(year, 4, 1), datetime.date(year, 6, 30)
    elif quarter == 3:
        return datetime.date(year, 7, 1), datetime.date(year, 9, 30)
    else:
        return datetime.date(year, 10, 1), datetime.date(year, 12, 31)

@login_required
def tb_trimestral_report(request):
    from .models import TBCase
    from datetime import date as date_cls
    from django.db.models import Q

    current_year = timezone.localdate().year
    current_quarter = (timezone.localdate().month - 1) // 3 + 1
    year = int(request.GET.get('year', current_year))
    quarter = int(request.GET.get('quarter', current_quarter))

    start_date, end_date = get_quarter_dates(year, quarter)

    quarter_map = {
        1: {'label': 'Q1: Janeiro - Marco',      'months': [1, 2, 3]},
        2: {'label': 'Q2: Abril - Junho',         'months': [4, 5, 6]},
        3: {'label': 'Q3: Julho - Setembro',      'months': [7, 8, 9]},
        4: {'label': 'Q4: Outobro - Dezembro',    'months': [10, 11, 12]},
    }
    qi = quarter_map.get(quarter, quarter_map[1])
    months = qi['months']

    qs = TBCase.objects.filter(date_started__year=year, date_started__month__in=months).select_related('patient')

    missing_hiv = qs.filter(Q(hiv_status__isnull=True) | Q(hiv_status='')).select_related('patient')
    missing_dm = qs.filter(Q(diabetes_status__isnull=True) | Q(diabetes_status='')).select_related('patient')

    FOUN = ['FOUN']; REKAIDA = ['RELAPSU']; RETRAT = ['DEPOIS_LAKON', 'FALHA']; SELUK = ['OUTRU']

    def cnt(q, pt=None, g=None):
        if pt: q = q.filter(patient_type__in=pt)
        if g: q = q.filter(patient__gender=g)
        return q.count()

    def b1_row(label, fk):
        q = qs.filter(**fk)
        return {'label':label, 'foun':cnt(q,FOUN), 'foun_m':cnt(q,FOUN,'M'), 'foun_f':cnt(q,FOUN,'F'), 'rekaida':cnt(q,REKAIDA), 'rekaida_m':cnt(q,REKAIDA,'M'), 'rekaida_f':cnt(q,REKAIDA,'F'), 'retrat':cnt(q,RETRAT), 'retrat_m':cnt(q,RETRAT,'M'), 'retrat_f':cnt(q,RETRAT,'F'), 'seluk':cnt(q,SELUK), 'seluk_m':cnt(q,SELUK,'M'), 'seluk_f':cnt(q,SELUK,'F'), 'total':q.count()}

    b1_rows = [
        b1_row('Pulmaun, bakteriolojiku konfirmadu',       {'classification': 'P',  'case_type': 'PTB_POS'}),
        b1_row('Pulmaun, diagnoza kliniku',                {'classification': 'P',  'case_type': 'PTB_NEG'}),
        b1_row('Extrapulmaun, bakteriolojiku konfirmadu',  {'classification': 'EP', 'case_type': 'PTB_POS'}),
        b1_row('Extrapulmonary, diagnoza kliniku',         {'classification': 'EP', 'case_type': 'EPTB'}),
    ]

    def _s(rows, f): return sum(r[f] for r in rows)
    b1 = {'rows':b1_rows, 'total_foun':_s(b1_rows,'foun'), 'total_foun_m':_s(b1_rows,'foun_m'), 'total_foun_f':_s(b1_rows,'foun_f'), 'total_rekaida':_s(b1_rows,'rekaida'), 'total_rekaida_m':_s(b1_rows,'rekaida_m'), 'total_rekaida_f':_s(b1_rows,'rekaida_f'), 'total_retrat':_s(b1_rows,'retrat'), 'total_retrat_m':_s(b1_rows,'retrat_m'), 'total_retrat_f':_s(b1_rows,'retrat_f'), 'total_seluk':_s(b1_rows,'seluk'), 'total_seluk_m':_s(b1_rows,'seluk_m'), 'total_seluk_f':_s(b1_rows,'seluk_f'), 'total_all':qs.count()}

    today = date_cls.today()
    AGS = [(0,4),(5,14),(15,24),(25,34),(35,44),(45,54),(55,64),(65,999)]

    def b2_row(label, fk):
        f = list(qs.filter(**fk).select_related('patient'))
        v = []; tm = tf = 0
        for lo, hi in AGS:
            mc = fc = 0
            for case in f:
                dob = case.patient.date_of_birth
                if not dob: continue
                age = (today - dob).days // 365
                if lo <= age <= hi:
                    if case.patient.gender == 'M': mc += 1
                    else: fc += 1
            tm += mc; tf += fc; v.append({'m': mc, 'f': fc})
        return {'label': label, 'age_groups': v, 'total_m': tm, 'total_f': tf, 'total': tm + tf}

    b2 = {'rows': [
        b2_row('Pulmaun, bakteriolojiku konfirmadu (F+R)', {'classification':'P', 'case_type':'PTB_POS','patient_type__in':FOUN+REKAIDA}),
        b2_row('Pulmaun, diagnoza kliniku (F+R)',          {'classification':'P', 'case_type':'PTB_NEG','patient_type__in':FOUN+REKAIDA}),
        b2_row('Extrapulmaun, bakteriolojiku konf. (F+R)', {'classification':'EP','case_type':'PTB_POS','patient_type__in':FOUN+REKAIDA}),
        b2_row('Extrapulmonary, diagnoza kliniku (F+R)',   {'classification':'EP','case_type':'EPTB',   'patient_type__in':FOUN+REKAIDA}),
        b2_row('Re-Tratamentu (La Inklui Rekaida)',        {'patient_type__in': RETRAT}),
        b2_row('Seluk',                                    {'patient_type__in': SELUK}),
    ]}

    sm_ex = gx_ex = sm_po = gx_po = 0
    try:
        from laboratory.models import LabResult
        lr_all = LabResult.objects.filter(completed_at__date__range=[start_date, end_date], lab_request__status='COMPLETED')
        afb_r = lr_all.filter(Q(lab_request__tests__name__icontains='AFB') | Q(lab_request__tests__name__icontains='ZN')).distinct()
        gx_r = lr_all.filter(lab_request__tests__name__icontains='GeneXpert').distinct()
        sm_ex = afb_r.count(); gx_ex = gx_r.count()
        sm_po = sum(1 for r in afb_r if r.result_data and '+' in str(r.result_data.get('microbiology',{}).get('zn_afb','')))
        gx_po = sum(1 for r in gx_r if r.result_data and 'pos' in str(r.result_data).lower())
    except: pass

    nr_qs = qs.filter(patient_type__in=FOUN+REKAIDA)
    gx_t = nr_qs.filter(initial_sputum__icontains='xpert').count()
    b3 = {'smear_exam': sm_ex, 'genexpert_exam': gx_ex, 'smear_pos': sm_po, 'genexpert_pos': gx_po, 'gx_total_tested': gx_t, 'gx_total_not_tested': max(0, nr_qs.count() - gx_t)}

    hiv_all = qs.exclude(hiv_status__isnull=True).exclude(hiv_status='').count()
    hiv_nr  = nr_qs.exclude(hiv_status__isnull=True).exclude(hiv_status='').count()
    hiv_p = qs.filter(hiv_status__icontains='reactive').exclude(hiv_status__icontains='non').count()
    art = sum(1 for c in qs if c.card_data.get('h_a_d'))
    cpt = sum(1 for c in qs if c.card_data.get('h_c_d'))
    b4 = {'hiv_tested_all': hiv_all, 'hiv_tested_new_relapse': hiv_nr, 'hiv_positive': hiv_p, 'hiv_art': art, 'hiv_cpt': cpt}

    fbs = qs.exclude(diabetes_status__isnull=True).exclude(diabetes_status='').count()
    dm = qs.filter(diabetes_status__icontains='dm').count()
    b5 = {'fbs_tested': fbs, 'dm_positive': dm}

    cu2 = c214 = ad = pl = 0
    for case in qs:
        cd = case.card_data or {}
        try: cu2 += int(cd.get('ltbi_child_u2',0) or 0); c214 += int(cd.get('ltbi_child_2to14',0) or 0); ad += int(cd.get('ltbi_adult',0) or 0); pl += int(cd.get('ltbi_plhiv',0) or 0)
        except: pass
    b6 = {'child_under2': cu2, 'child_2to14': c214, 'adult': ad, 'plhiv': pl}

    return render(request, 'medical_records/tb_trimestral_report.html', {
        'year': year, 'quarter': quarter, 'quarter_label': qi['label'], 'year_range': range(2020, current_year + 2),
        'b1': b1, 'b2': b2, 'b3': b3, 'b4': b4, 'b5': b5, 'b6': b6,
        'missing_hiv': missing_hiv, 'missing_dm': missing_dm,
    })

@login_required
def tb_trimestral_export_excel(request):
    import io
    from datetime import date as date_cls
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse('openpyxl not installed', status=500)

    from django.http import HttpResponse
    from .models import TBCase
    from django.db.models import Q

    year = int(request.GET.get('year', timezone.localdate().year))
    quarter = int(request.GET.get('quarter', (timezone.localdate().month - 1) // 3 + 1))
    sd, ed = get_quarter_dates(year, quarter)

    quarter_map = {1: {'label': 'Q1: Janeiro - Marco', 'months': [1,2,3]}, 2: {'label': 'Q2: Abril - Junho', 'months': [4,5,6]}, 3: {'label': 'Q3: Julho - Setembro', 'months': [7,8,9]}, 4: {'label': 'Q4: Outobro - Dezembro', 'months': [10,11,12]}}
    qi = quarter_map.get(quarter, quarter_map[1])
    qs = TBCase.objects.filter(date_started__year=year, date_started__month__in=qi['months']).select_related('patient')

    FOUN = ['FOUN']; REKAIDA = ['RELAPSU']; RETRAT = ['DEPOIS_LAKON', 'FALHA']; SELUK = ['OUTRU']
    def cnt(q, pt=None, g=None):
        if pt: q = q.filter(patient_type__in=pt)
        if g: q = q.filter(patient__gender=g)
        return q.count()

    b1_rows = []
    for lbl, fk in [('Pulmaun, bakteriolojiku konfirmadu', {'classification':'P', 'case_type':'PTB_POS'}), ('Pulmaun, diagnoza kliniku', {'classification':'P', 'case_type':'PTB_NEG'}), ('Extrapulmaun, bakteriolojiku konf.', {'classification':'EP', 'case_type':'PTB_POS'}), ('Extrapulmonary, diagnoza kliniku', {'classification':'EP', 'case_type':'EPTB'})]:
        q = qs.filter(**fk)
        b1_rows.append([lbl, cnt(q,FOUN), cnt(q,FOUN,'M'), cnt(q,FOUN,'F'), cnt(q,REKAIDA), cnt(q,REKAIDA,'M'), cnt(q,REKAIDA,'F'), cnt(q,RETRAT), cnt(q,RETRAT,'M'), cnt(q,RETRAT,'F'), cnt(q,SELUK), cnt(q,SELUK,'M'), cnt(q,SELUK,'F'), q.count()])

    today = date_cls.today()
    AGS = [(0,4),(5,14),(15,24),(25,34),(35,44),(45,54),(55,64),(65,999)]
    b2_rows = []
    for lbl, fk in [('Pulmaun, bakt. konf. (F+R)', {'classification':'P', 'case_type':'PTB_POS','patient_type__in':FOUN+REKAIDA}), ('Pulmaun, diag. klin. (F+R)', {'classification':'P', 'case_type':'PTB_NEG','patient_type__in':FOUN+REKAIDA}), ('Extrapulmaun, bakt. konf. (F+R)', {'classification':'EP','case_type':'PTB_POS','patient_type__in':FOUN+REKAIDA}), ('Extrapulmonary, diag. klin. (F+R)', {'classification':'EP','case_type':'EPTB', 'patient_type__in':FOUN+REKAIDA}), ('Re-Tratamentu (la rekaida)', {'patient_type__in':RETRAT}), ('Seluk', {'patient_type__in':SELUK})]:
        f = list(qs.filter(**fk).select_related('patient'))
        v = [lbl]; tm = tf = 0
        for lo, hi in AGS:
            m = fc = 0
            for c in f:
                dob = c.patient.date_of_birth
                if dob and lo <= (today - dob).days // 365 <= hi:
                    if c.patient.gender == 'M': m += 1
                    else: fc += 1
            tm += m; tf += fc; v += [m, fc]
        v += [tm, tf, tm+tf]; b2_rows.append(v)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f'TB_Q{quarter}_{year}'
    f_hdr = Font(bold=True, size=10); f_blk = Font(bold=True, size=9, color='FFFFFF'); fill_blk = PatternFill('solid', fgColor='1565C0'); fill_sub = PatternFill('solid', fgColor='BBDEFB'); fill_tot = PatternFill('solid', fgColor='FFF8E1'); border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for i, t in enumerate(['REPUBLICA DEMOCRATICA DE TIMOR LESTE','MINISTÉRIO DA SAÚDE','SERVIÇO SAÚDE MUNICIPIO DE DILI','PROGRAMA NASIONAL KONTROLA TUBERKULOZU'], 1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=15)
        c = ws.cell(i, 1, t); c.font = f_hdr; c.alignment = Alignment(horizontal='center')

    ws.merge_cells(5,1,5,15); c = ws.cell(5,1, "RELATORIO TRIMESTRAL DETEKSAUN KAZU TUBERKULOZU"); c.font = Font(bold=True, size=12, underline='single'); c.alignment = Alignment(horizontal='center')
    
    r = 12; ws.merge_cells(r,1,r,14); c = ws.cell(r,1, "Block 1 : Kazu TB hotu ne'ebe rejistu"); c.font=f_blk; c.fill=fill_blk
    r += 1; hdrs1 = ['Tipu Kazu','Foun Total','M','F','Rekaida Total','M','F','Re-Tratamentu Total','M','F','Seluk Total','M','F','TOTAL']
    for i, h in enumerate(hdrs1, 1): c=ws.cell(r,i,h); c.font=Font(bold=True, size=8); c.fill=fill_sub; c.border=border

    for row_data in b1_rows:
        r += 1
        for i, val in enumerate(row_data, 1): c=ws.cell(r,i,val); c.border=border; c.font=Font(size=8)
        ws.cell(r,14).fill=fill_tot

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="TB_Report_Q{quarter}_{year}.xlsx"'
    return response
